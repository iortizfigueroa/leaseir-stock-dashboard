#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stock_por_coste.py
==================

Genera Stock_por_coste.xlsx con UNA tabla plana ordenada por coste unitario DESC.
Cada fila: SPEC | Descripcion | Proveedor | Tipo | Coste/u | U.raw | U.emb | U.total | Valor

Aplica:
- refine_type (clasificacion fina)
- Inferencia de coste para SPECs sin stock raw a partir de companeros (misma
  descripcion normalizada quitando voltajes/sufijos) o coste teorico del BOM.
"""

from __future__ import annotations
import sys
import re
import importlib.util
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(r"/sessions/stoic-gallant-albattani/mnt/Donet (inventario)")
DAILY_FILE = PROJECT_DIR / "ejercicio 28-05-2026.xlsx"
OUTPUT_FILE = PROJECT_DIR / "Stock_por_coste.xlsx"

sys.path.insert(0, str(PROJECT_DIR))
if "master" in sys.modules:
    master = sys.modules["master"]
else:
    spec_m = importlib.util.spec_from_file_location("master", PROJECT_DIR / "inventario_master.py")
    master = importlib.util.module_from_spec(spec_m)
    spec_m.loader.exec_module(master)
    sys.modules["master"] = master
if not master.ESCANDALLOS:
    master.load_escandallos()
if "prov" in sys.modules:
    prov = sys.modules["prov"]
else:
    spec_p = importlib.util.spec_from_file_location("p", PROJECT_DIR / "stock_por_proveedor.py")
    prov = importlib.util.module_from_spec(spec_p)
    spec_p.loader.exec_module(prov)
    sys.modules["prov"] = prov


def normalize_desc(desc: str) -> str:
    """Normaliza la descripcion para encontrar 'companeros' (mismo SPEC con
    voltaje/idioma/variante diferente)."""
    d = (desc or "")
    # Quitar voltajes
    d = re.sub(r"\b(110|115|120|220|230|240)\s*[vV]\b", "", d)
    # Quitar idiomas en parentesis
    d = re.sub(r"\((english|spanish|italian|portuguese|french|german)\)", "", d, flags=re.IGNORECASE)
    # Quitar refs de catalogo entre []
    d = re.sub(r"\[[^\]]+\]", "", d)
    # Quitar guiones y signos
    d = re.sub(r"[\-_]", " ", d)
    # Espacios
    d = re.sub(r"\s+", " ", d).strip().lower()
    return d


def main():
    print("[1/5] Cargando...")
    master.load_escandallos()
    suppliers = prov.load_suppliers_from_xlsx(DAILY_FILE)
    inventory = prov.load_inventory_from_xlsx(DAILY_FILE)
    of_rows = prov.load_ofs_from_xlsx(DAILY_FILE, inventory)
    inventory.extend(of_rows)
    for row in inventory:
        row["cat"] = master.classify(row, suppliers)
    unit_cost = master.build_unit_costs(inventory)
    recon = master.reconcile(inventory, unit_cost)
    rollup_rows = master.build_rollup(inventory, recon, unit_cost, suppliers, {})

    # Reclasificacion fina + inferir proveedor por descripcion cuando falta
    for r in rollup_rows:
        # Fallback proveedor para SPECs claramente Monocrom sin mapeo
        if not r.get("supplier"):
            d = r.get("desc", "") or ""
            if re.search(r"laser diode|\bENSPSTM\b|ensptm|\bLBS-|\bmonocrom\b", d, re.IGNORECASE):
                r["supplier"] = "MONOCROM S.L"
        r["tipo"] = prov.refine_type(r["tipo"], r.get("supplier", ""),
                                      r.get("desc", ""), r.get("unit_cost", 0))

    print("[2/5] Construyendo mapping de costes por descripcion normalizada...")
    # Para cada SPEC con coste > 0, registrar su coste bajo la descripcion normalizada
    cost_by_norm: dict[str, list] = defaultdict(list)
    for r in rollup_rows:
        if r["unit_cost"] and r["unit_cost"] > 0:
            nd = normalize_desc(r["desc"])
            if nd:
                cost_by_norm[nd].append(r["unit_cost"])

    # Coste medio por descripcion normalizada (puede haber multiples SPECs con
    # misma normalizacion: ej Hailea 230V y 110V)
    avg_cost_by_norm = {nd: sum(costs) / len(costs) for nd, costs in cost_by_norm.items()}

    print("[3/5] Inferir coste para SPECs sin stock raw...")
    # Tambien coste material teorico del BOM
    def cost_material_teorico(canon_spec):
        if canon_spec not in master.ESCANDALLOS:
            return 0.0
        exp = master.expand_bom(canon_spec)
        total = 0.0
        for raw_canon, q in exp.items():
            uc = unit_cost.get(raw_canon) or unit_cost.get(master.base_spec(raw_canon)) or 0.0
            total += q * uc
        return total

    n_inferidos = 0
    for r in rollup_rows:
        if r["unit_cost"] and r["unit_cost"] > 0:
            r["cost_source"] = "inventario"
            continue
        # PRIMERO: division directa si val_total > 0 y total > 0
        if (r["val_total"] or 0) > 0 and (r["total"] or 0) > 0:
            r["unit_cost"] = r["val_total"] / r["total"]
            r["cost_source"] = "val/qty"
            n_inferidos += 1
            continue
        # Companero (misma desc normalizada)
        nd = normalize_desc(r["desc"])
        if nd and nd in avg_cost_by_norm:
            r["unit_cost"] = avg_cost_by_norm[nd]
            r["cost_source"] = "companero"
            n_inferidos += 1
            continue
        # Coste teorico via BOM
        tc = cost_material_teorico(r["spec"])
        if tc > 0:
            r["unit_cost"] = tc
            r["cost_source"] = "BOM teorico"
            n_inferidos += 1
            continue
        r["cost_source"] = "(sin coste)"

    print(f"      {n_inferidos} SPECs con coste inferido")

    print("[4/5] Filtrando y ordenando...")
    # Filtrar items con stock > 0 (raw o embebido)
    rows = [r for r in rollup_rows if (r["total"] or 0) > 0]
    rows.sort(key=lambda x: -(x["unit_cost"] or 0))

    print(f"      {len(rows)} SPECs con stock fisico")

    print("[5/5] Escribiendo Excel...")
    write_excel(rows, recon)
    print(f"GUARDADO: {OUTPUT_FILE}")


def write_excel(rows, recon):
    wb = openpyxl.Workbook()

    H_FILL = PatternFill("solid", fgColor="305496")
    H_FONT = Font(bold=True, color="FFFFFF")
    INV_FILL = PatternFill("solid", fgColor="E2EFDA")  # verde claro
    INFER_FILL = PatternFill("solid", fgColor="FFEB9C")  # amarillo
    BOM_FILL = PatternFill("solid", fgColor="DDEBF7")    # azul claro
    NONE_FILL = PatternFill("solid", fgColor="FFC7CE")   # rojo claro
    TOTAL_FILL = PatternFill("solid", fgColor="FFE699")

    EUR = '#,##0.00 "EUR"'
    INT = '#,##0'
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Por coste unitario"

    hdr = ["#", "SPEC", "Descripcion", "Proveedor", "Tipo",
           "Coste/u EUR", "Fuente coste", "U. RAW", "U. emb", "U. TOTAL", "Valor TOTAL"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    color_by_src = {
        "inventario": INV_FILL,
        "companero": INFER_FILL,
        "BOM teorico": BOM_FILL,
        "(sin coste)": NONE_FILL,
    }

    for i, r in enumerate(rows, start=2):
        src = r.get("cost_source", "(sin coste)")
        ws.cell(row=i, column=1, value=i-1).number_format = INT
        ws.cell(row=i, column=2, value=r["spec"])
        ws.cell(row=i, column=3, value=(r["desc"] or "")[:90])
        ws.cell(row=i, column=4, value=(r["supplier"] or "(SIN PROVEEDOR)"))
        ws.cell(row=i, column=5, value=r["tipo"])
        ws.cell(row=i, column=6, value=r["unit_cost"] or 0).number_format = EUR
        ws.cell(row=i, column=7, value=src)
        ws.cell(row=i, column=8, value=r["u_raw"]).number_format = INT
        ws.cell(row=i, column=9, value=r["u_emb"]).number_format = INT
        ws.cell(row=i, column=10, value=r["total"]).number_format = INT
        # Valor recalculado: si u_raw=0 pero hay coste inferido, valor = total * coste/u
        if r["val_total"] and r["val_total"] > 0:
            v = r["val_total"]
        else:
            v = (r["unit_cost"] or 0) * (r["total"] or 0)
        ws.cell(row=i, column=11, value=v).number_format = EUR
        # Color por fuente
        fill = color_by_src.get(src, NONE_FILL)
        for c in range(1, 12):
            ws.cell(row=i, column=c).fill = fill
            ws.cell(row=i, column=c).border = border

    widths = [5, 14, 60, 35, 28, 14, 16, 10, 10, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Pestania de leyenda y resumen ----
    ws2 = wb.create_sheet("Leyenda", 0)
    ws2["A1"] = "STOCK POR COSTE UNITARIO - al 27-05-2026"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Cada fila = 1 SPEC. Ordenado por coste/u DESCENDENTE."
    ws2["A4"] = "El color de la fila indica de donde viene el coste:"

    ws2["A6"].fill = INV_FILL; ws2["B6"] = "Coste medio del inventario (qty raw > 0)"
    ws2["A7"].fill = INFER_FILL; ws2["B7"] = "Inferido de companero (misma desc, voltaje distinto)"
    ws2["A8"].fill = BOM_FILL; ws2["B8"] = "Coste teorico del BOM (suma de raws del escandallo)"
    ws2["A9"].fill = NONE_FILL; ws2["B9"] = "Sin coste (valoran 0 EUR)"

    ws2["A11"] = "Resumen de las 4 bolsas:"
    ws2["A11"].font = Font(bold=True)
    ws2.cell(row=12, column=1, value="A - RAW + DEFECTIVE")
    ws2.cell(row=12, column=2, value=recon["bolsa_A_val"]).number_format = EUR
    ws2.cell(row=13, column=1, value="B - Material embebido")
    ws2.cell(row=13, column=2, value=recon["bolsa_B_val"]).number_format = EUR
    ws2.cell(row=14, column=1, value="C - Overhead / drift")
    ws2.cell(row=14, column=2, value=recon["bolsa_C_val"]).number_format = EUR
    ws2.cell(row=15, column=1, value="D - WIPs sin BOM")
    ws2.cell(row=15, column=2, value=recon["bolsa_D_val"]).number_format = EUR
    total_val = sum(recon[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
    ws2.cell(row=16, column=2, value=total_val).number_format = EUR
    ws2.cell(row=16, column=2).font = Font(bold=True)
    ws2.cell(row=16, column=2).fill = TOTAL_FILL

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 60

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    main()
