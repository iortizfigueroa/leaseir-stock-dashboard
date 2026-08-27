#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
vision_html_full.py
===================

Genera Vision_Stock_Mes.html: dashboard interactivo con la evolucion DIARIA
del stock por SPEC desde el 30-abril hasta el 29-mayo (todos los dias habiles).

- Dia 30-abr: inventario real con bolsas A+B+C+D + fila "GAP NO EXPLICADO" para
  que el inferido del 29-may cuadre con el real.
- Dias intermedios (4-may a 26-may): stock calculado aplicando entradas/salidas
  del dia sobre el dia anterior.
- Dias 27/28/29-may: inventario real (tomado del fichero ejercicio del dia).

Tabs: Stock evolutivo, Entradas (todo el mes), Salidas (todo el mes), Salidas
por cliente (todo el mes).
"""

from __future__ import annotations
import sys
import json
import re
import importlib.util
from collections import defaultdict
from datetime import datetime, date as _date
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_HTML = PROJECT_DIR / "Vision_Stock_Mes.html"
TEMPLATE = PROJECT_DIR / "vision_html_template_full.html"
# v2 (Vercel): mismo payload, template con Alertas + Prevision. Solo se genera si existe.
OUTPUT_HTML_V2 = PROJECT_DIR / "Vision_Stock_Mes_v2.html"
TEMPLATE_V2 = PROJECT_DIR / "vision_html_template_v2.html"
COST_THRESHOLD = 50.0

sys.path.insert(0, str(PROJECT_DIR))

spec_m = importlib.util.spec_from_file_location("evol", PROJECT_DIR / "actualizar_evolutivo.py")
evol = importlib.util.module_from_spec(spec_m)
spec_m.loader.exec_module(evol)
master = evol.master
prov = evol.prov

# Reusar funciones del reconcile (loader robust + breakdown)
spec_r = importlib.util.spec_from_file_location("rec", PROJECT_DIR / "reconcile_abril.py")
rec = importlib.util.module_from_spec(spec_r)
spec_r.loader.exec_module(rec)


def expand_bom_ui(code, *args, **kwargs):
    """expand_bom para la capa visual del dashboard: filtra las lineas
    sinteticas de proceso externo (PROC-*). El proceso se crea al PRODUCIR
    (sin fila de movimiento), asi que si viajara con las entradas/salidas la
    propagacion hacia atras acumularia stock de proceso inexistente en el
    pasado. En la capa visual su stock queda constante al nivel del ancla;
    la reconciliacion (bolsas A/B/C/D) si lo usa completo via master."""
    return {r: q for r, q in master.expand_bom(code, *args, **kwargs).items()
            if not str(r).startswith("PROC-")}


def normalize_code(code):
    c = str(code or "").strip()
    if not c:
        return ""
    c = re.sub(r"_C$", "_c", c)
    return master.canonical(c)


def load_ofs_en_vuelo_detalle(daily_file):
    """Lee la hoja 'en vuelo' y devuelve {of_num_str: {spec, desc, status,
    emitido, recibido, pendiente}} para la FASE 3 (flujo de fabrica)."""
    out = {}
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True, read_only=True)
    except Exception:
        return out
    try:
        sn = next((s for s in wb.sheetnames if 'vuelo' in s.lower()), None)
        if not sn:
            return out
        for r in wb[sn].iter_rows(values_only=True, min_row=2):
            if not r or r[0] is None or len(r) < 8:
                continue
            of_num = str(r[0]).strip()
            spec_code = (str(r[2]) if r[2] is not None else "").strip()
            if not of_num or not spec_code:
                continue
            desc = (str(r[3]) if r[3] is not None else "").strip()
            status = (str(r[1]) if r[1] is not None else "").strip()
            def _f(x):
                try:
                    return float(x or 0)
                except (TypeError, ValueError):
                    return 0.0
            out[of_num] = {
                "spec": spec_code, "desc": desc[:60], "status": status,
                "emitido": _f(r[5]), "recibido": _f(r[6]), "pendiente": _f(r[7]),
            }
    finally:
        wb.close()
    return out


def load_movimientos_por_dia(daily_file):
    """Devuelve dict {date: {'compras': {spec: qty}, 'salidas': {spec: qty}}}"""
    by_day = defaultdict(lambda: {"compras": defaultdict(float), "salidas": defaultdict(float)})
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True)
    except Exception:
        return by_day

    def get_date(f):
        if isinstance(f, datetime):
            return f.date()
        if isinstance(f, _date):
            return f
        return None

    for sn in wb.sheetnames:
        if 'entradas' in sn.lower() and 'spec' in sn.lower():
            sh = wb[sn]
            hdrs = [str(c.value or '').lower() for c in sh[1]]
            fecha_c = next((i for i, h in enumerate(hdrs) if 'fecha' in h and 'albar' in h), None)
            if fecha_c is None:
                fecha_c = next((i for i, h in enumerate(hdrs) if 'fecha' in h), None)
            spec_c = next((i for i, h in enumerate(hdrs) if 'artículo' in h or 'articulo' in h or 'código' in h), None)
            qty_c = next((i for i, h in enumerate(hdrs) if 'cantidad' in h), None)
            if fecha_c is None or spec_c is None or qty_c is None:
                continue
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) <= max(fecha_c, spec_c, qty_c):
                    continue
                d = get_date(row[fecha_c])
                if d is None:
                    continue
                code = normalize_code(row[spec_c])
                if not code:
                    continue
                try:
                    qty = float(row[qty_c] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in expand_bom_ui(code).items():
                        by_day[d]["compras"][raw] += q * qty
                else:
                    by_day[d]["compras"][code] += qty
            break

    for sn in wb.sheetnames:
        snl = sn.lower()
        if 'entregas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 8:
                    continue
                d = get_date(row[0])
                if d is None:
                    continue
                code = normalize_code(row[5])
                if not code:
                    continue
                try:
                    qty = float(row[7] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in expand_bom_ui(code).items():
                        by_day[d]["salidas"][raw] += q * qty
                else:
                    by_day[d]["salidas"][code] += qty
        if 'salidas directas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 5:
                    continue
                d = get_date(row[0])
                if d is None:
                    continue
                raw_code = str(row[2] or '').strip()
                if not raw_code or 'artículo' in raw_code.lower():
                    continue
                code = normalize_code(raw_code)
                try:
                    qty = float(row[4] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in expand_bom_ui(code).items():
                        by_day[d]["salidas"][raw] += q * qty
                else:
                    by_day[d]["salidas"][code] += qty
    return by_day


def cargar_salidas_con_cliente_full(daily_file):
    """Como rec.cargar_salidas_con_cliente pero SIN filtrar por dia (todo el rango)."""
    from datetime import datetime as _dt, date as _dt2
    rows = []
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True)
    except Exception:
        return rows

    def gd(f):
        if isinstance(f, _dt): return f.date()
        if isinstance(f, _dt2): return f
        return None

    for sn in wb.sheetnames:
        snl = sn.lower()
        if 'entregas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 10:
                    continue
                d = gd(row[0])
                if d is None: continue
                code = str(row[5] or '').strip()
                if not code: continue
                try: qty = float(row[7] or 0)
                except: continue
                if qty <= 0: continue
                try: coste = float(row[9] or 0)
                except: coste = 0
                rows.append({
                    "fecha": d.isoformat(),
                    "spec_emitido": code,
                    "spec_canon": normalize_code(code),
                    "desc": str(row[6] or '')[:80],
                    "cliente": (str(row[4] or '').strip() or "(SIN CLIENTE)"),
                    "qty": qty, "coste_unit": coste, "valor": coste * qty,
                    "fuente": "entrega",
                    "lote": str(row[8] or '').strip(),
                    "doc": str(row[2] or '').strip(),
                })
        if 'salidas directas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 5: continue
                d = gd(row[0])
                if d is None: continue
                raw_code = str(row[2] or '').strip()
                if not raw_code or 'artículo' in raw_code.lower(): continue
                try: qty = float(row[4] or 0)
                except: continue
                if qty <= 0: continue
                try: coste = float(row[5] or 0) if len(row) > 5 else 0
                except: coste = 0
                rows.append({
                    "fecha": d.isoformat(),
                    "spec_emitido": raw_code,
                    "spec_canon": normalize_code(raw_code),
                    "desc": str(row[3] or '')[:80],
                    "cliente": "(interno / consumo SAT)",
                    "qty": qty, "coste_unit": coste, "valor": coste * qty,
                    "fuente": "directa",
                    "lote": "",
                    "doc": str(row[1] or '').strip(),
                })
    return rows


def cargar_entradas_full(daily_file):
    """Todas las entradas del rango con SPEC."""
    from datetime import datetime as _dt, date as _dt2
    rows = []
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True)
    except Exception:
        return rows
    def gd(f):
        if isinstance(f, _dt): return f.date()
        if isinstance(f, _dt2): return f
        return None
    for sn in wb.sheetnames:
        if 'entradas' in sn.lower() and 'spec' in sn.lower():
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 13: continue
                d = gd(row[6])
                if d is None: continue
                code = str(row[8] or '').strip()
                if not code: continue
                try: qty = float(row[10] or 0)
                except: continue
                if qty <= 0: continue
                try: cu = float(row[11] or 0)
                except: cu = 0
                try: val = float(row[12] or 0)
                except: val = cu * qty
                rows.append({
                    "fecha": d.isoformat(),
                    "spec_emitido": code,
                    "spec_canon": normalize_code(code),
                    "desc": str(row[9] or '')[:80],
                    "proveedor": str(row[1] or '').strip(),
                    "qty": qty, "coste_unit": cu, "valor": val,
                    "doc": str(row[3] or '').strip(),
                })
            break
    return rows



# ============================================================
# SIMULACIÓN — leer Excel Live y calcular forecast stock
# ============================================================

MESES_SIM = ["Jun-26", "Jul-26", "Ago-26", "Sep-26", "Oct-26", "Nov-26", "Dic-26"]


def read_live_excel(live_path):
    """Lee forecast, mínimos y compras obligatorias del Excel Live."""
    if not live_path.exists():
        return None, None, None
    try:
        wb = openpyxl.load_workbook(live_path, data_only=True)
    except Exception:
        return None, None, None

    forecast = {}
    ws = wb["1. Forecast (manual)"]
    for r in range(6, ws.max_row + 1):
        sp = ws.cell(row=r, column=1).value
        if not sp: continue
        sp = str(sp).strip()
        for i, mes in enumerate(MESES_SIM + ["Ene-27","Feb-27","Mar-27","Abr-27","May-27"]):
            v = ws.cell(row=r, column=4 + i).value
            if v is not None and v != "":
                try: forecast.setdefault(sp, {})[mes] = float(v)
                except: pass

    minimos = {}
    ws = wb["2. Mínimos (manual)"]
    for r in range(6, ws.max_row + 1):
        sp = ws.cell(row=r, column=1).value
        if not sp: continue
        sp = str(sp).strip()
        v6 = ws.cell(row=r, column=6).value
        v7 = ws.cell(row=r, column=7).value
        min_val = v7 if (v7 is not None and v7 != "") else v6
        if min_val is not None and min_val != "":
            try: minimos[sp] = float(min_val)
            except: pass

    compras_obl = []
    ws = wb["3. Compras oblig (manual)"]
    for r in range(7, ws.max_row + 1):
        sp = ws.cell(row=r, column=1).value
        prov_ = ws.cell(row=r, column=3).value
        uds = ws.cell(row=r, column=4).value
        importe = ws.cell(row=r, column=5).value
        mes = ws.cell(row=r, column=6).value
        if (sp or prov_) and mes:
            compras_obl.append({
                "spec": (str(sp).strip() if sp else None),
                "proveedor": (str(prov_).strip() if prov_ else None),
                "uds": float(uds) if (uds not in (None, "")) else None,
                "importe": float(importe) if (importe not in (None, "")) else None,
                "mes": str(mes).strip(),
            })
    return forecast, minimos, compras_obl


def compute_simulation(stock_real_29may, forecast, minimos, compras_obl):
    """Devuelve {spec: {mes: {stock_fin, compra, cu, desc, sup, tipo, min}}}"""
    from collections import defaultdict

    # Demanda por raw via BOM
    demand = defaultdict(lambda: defaultdict(float))
    for wip, meses_dict in (forecast or {}).items():
        canon = master.canonical(wip)
        if canon not in master.ESCANDALLOS:
            canon_c = canon + "_c"
            if canon_c in master.ESCANDALLOS: canon = canon_c
        if canon not in master.ESCANDALLOS:
            continue
        try: exp = expand_bom_ui(canon)
        except: continue
        for mes, uds in meses_dict.items():
            if uds <= 0: continue
            for raw_canon, q in exp.items():
                raw_base = master.base_spec(raw_canon)
                demand[raw_base][mes] += q * uds

    # Compras obligatorias por SPEC
    cobl = defaultdict(lambda: defaultdict(float))
    for c in (compras_obl or []):
        if c["spec"] and c["uds"]:
            cobl[c["spec"]][c["mes"]] += c["uds"]

    # SPECs relevantes
    all_specs = set(stock_real_29may.keys()) | set(demand.keys()) | set(minimos.keys()) | set(cobl.keys())

    sim = {}
    for sp in all_specs:
        info = stock_real_29may.get(sp, {"u_tot": 0, "cu": 0, "desc": "", "supplier": "", "tipo": ""})
        stock_act = info.get("u_tot", 0)
        cu = info.get("cu", 0)
        min_manual = (minimos or {}).get(sp)
        if min_manual is None:
            min_manual = demand[sp].get("Jun-26", 0)
        per_month = {}
        for mes in MESES_SIM:
            d = demand[sp].get(mes, 0)
            c_obl = cobl[sp].get(mes, 0)
            stock_fin = stock_act + c_obl - d
            compra_extra = 0
            if stock_fin < min_manual:
                compra_extra = min_manual - stock_fin
                stock_fin += compra_extra
            per_month[mes] = {
                "stock_fin": stock_fin,
                "compra": c_obl + compra_extra,
                "demanda": d,
            }
            stock_act = stock_fin
        # Solo incluir si hay movimiento (demand o compra)
        total_movement = sum(per_month[m]["compra"] + per_month[m]["demanda"] for m in MESES_SIM)
        if total_movement > 0 or sp in (minimos or {}):
            sim[sp] = {
                "spec": sp,
                "desc": info.get("desc", ""),
                "sup": info.get("supplier", ""),
                "tipo": info.get("tipo", ""),
                "cu": cu,
                "stock_29may": info.get("u_tot", 0),  # nombre legacy: ahora es stock del anchor (01-jun)
                "minimo": min_manual,
                "months": per_month,
            }
    return sim


def main():
    print("[1/5] Cargando 30-abril (origen real)...")
    info_30abr, recon_30abr = rec.build_totales_por_spec(rec.ABRIL_FILE)
    total_30abr_real = sum(recon_30abr[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
    print(f"  30-abr REAL = {total_30abr_real:,.2f} EUR")

    # AUTO-DETECT ejercicios disponibles (ordenados cronológicamente).
    # Cualquier `ejercicio DD-MM-YYYY.xlsx` en PROJECT_DIR se incluye automáticamente.
    import re as _re
    ej_files = []
    for p in PROJECT_DIR.glob("ejercicio *.xlsx"):
        m = _re.match(r"ejercicio (\d{2})-(\d{2})-(\d{4})\.xlsx", p.name)
        if m:
            dd, mm, yyyy = m.groups()
            lbl = f"{dd}-{mm}"
            iso = f"{yyyy}-{mm}-{dd}"
            ej_files.append((iso, lbl, p))
    ej_files.sort()  # cronológico
    if not ej_files:
        sys.exit("ERROR: no se encontró ningún ejercicio*.xlsx en PROJECT_DIR")

    print(f"[2/5] Cargando {len(ej_files)} ejercicios disponibles...")
    days_real = {}
    for iso, lbl, fn in ej_files:
        info, recon = rec.build_totales_por_spec(fn)
        days_real[lbl] = {"info": info, "recon": recon,
                          "total": sum(recon[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))}
        print(f"  {lbl} REAL = {days_real[lbl]['total']:,.2f} EUR")

    # El ANCHOR es el ejercicio más reciente
    ANCHOR_LBL = ej_files[-1][1]
    print(f"[3/5] Cargando movimientos diarios (anchor = {ANCHOR_LBL})...")

    # === FIX PROVEEDORES GLOBAL ===
    # La hoja "proveedores" del ejercicio contiene el supplier de TODOS los SPECs.
    # El loader auxiliar busca "proveedores-SPEC" (nombre antiguo) y no encuentra nada.
    # Cargamos aqui la hoja real y sobrescribimos supplier en todos los info de days_real.
    _global_suppliers = {}
    try:
        _wb_g = openpyxl.load_workbook(ej_files[-1][2], data_only=True, read_only=True)
        _sh_g_name = next((n for n in _wb_g.sheetnames if n.lower().startswith("proveedor")), None)
        if _sh_g_name:
            _sh_g = _wb_g[_sh_g_name]
            _agg = defaultdict(set)
            for _r in _sh_g.iter_rows(values_only=True, min_row=2):
                if not _r or not _r[0]:
                    continue
                _code = str(_r[0]).strip()
                _sup = str((_r[3] if len(_r) > 3 else "") or "").strip()
                if not _sup:
                    continue
                _agg[normalize_code(_code)].add(_sup)
                _agg[_code].add(_sup)
            for _k, _names in _agg.items():
                _global_suppliers[_k] = next(iter(_names)) if len(_names) == 1 else " / ".join(sorted(_names))
        _wb_g.close()
        print(f"  Proveedores cargados de hoja '{_sh_g_name}': {len(_global_suppliers)} entries")
    except Exception as _e:
        print(f"  WARN suppliers: {_e}")

    # Sobrescribir supplier en info de cada dia (afecta a Stock, Entradas, Salidas, Cliente)
    for _lbl_x, _d_x in days_real.items():
        for _sp_x, _i_x in _d_x["info"].items():
            _cur_x = (_i_x.get("supplier", "") or "").strip()
            if _cur_x and _cur_x not in ("(SIN PROVEEDOR)", "(contable)"):
                continue
            _new_x = (_global_suppliers.get(_sp_x)
                      or _global_suppliers.get(normalize_code(_sp_x))
                      or _global_suppliers.get(master.base_spec(normalize_code(_sp_x))))
            if _new_x:
                _i_x["supplier"] = _new_x


    # Cargar movimientos de TODOS los ejercicios (el archivo de cada día puede traer todos los movimientos previos)
    mov_by_day = defaultdict(lambda: {"compras": defaultdict(float), "salidas": defaultdict(float)})
    for iso, lbl, fn in ej_files:
        mov_extra = load_movimientos_por_dia(fn)
        for d, m in mov_extra.items():
            if d not in mov_by_day:
                mov_by_day[d] = {"compras": defaultdict(float), "salidas": defaultdict(float)}
            for sp, q in m["compras"].items():
                mov_by_day[d]["compras"][sp] = max(mov_by_day[d]["compras"][sp], q)  # dedupe
            for sp, q in m["salidas"].items():
                mov_by_day[d]["salidas"][sp] = max(mov_by_day[d]["salidas"][sp], q)
    sorted_days = sorted(mov_by_day.keys())
    print(f"  {len(sorted_days)} dias con movimientos: {sorted_days[0]} a {sorted_days[-1]}")

    # Entradas y salidas full. Dedupe ENTRE ficheros (el mismo movimiento se repite en
    # varios ejercicios) pero CONSERVANDO las lineas repetidas REALES dentro de un mismo
    # fichero: un albaran de N equipos identicos trae N lineas iguales (1 ud por numero
    # de serie) y el dedupe plano se las comia (~479k EUR de salidas perdidas).
    # Estrategia: por cada clave, quedarnos con el MAXIMO de repeticiones visto en un
    # mismo fichero (igual que hace mov_by_day con las cantidades).
    def _dedupe_max_por_fichero(rows_por_fichero, keyf):
        best = {}
        for rows_f in rows_por_fichero:
            local = defaultdict(list)
            for r in rows_f:
                local[keyf(r)].append(r)
            for k, rs in local.items():
                if k not in best or len(rs) > len(best[k]):
                    best[k] = rs
        out = []
        for rs in best.values():
            out.extend(rs)
        return out

    entradas_por_fichero = []
    salidas_por_fichero = []
    for iso, lbl, fn in ej_files:
        entradas_por_fichero.append(cargar_entradas_full(fn))
        salidas_por_fichero.append(cargar_salidas_con_cliente_full(fn))
    entradas_full = _dedupe_max_por_fichero(
        entradas_por_fichero, lambda e: (e["fecha"], e["spec_emitido"], e["doc"], e["qty"]))
    salidas_full = _dedupe_max_por_fichero(
        salidas_por_fichero, lambda s: (s["fecha"], s["spec_emitido"], s["doc"], s["qty"], s["cliente"]))

    print("[4/5] Reconstruyendo stock dia a dia (BACKWARD desde 01-jun)...")
    # Lista de etiquetas en orden: 30-04, luego dias con movimientos
    day_labels = ["30-04"] + [d.strftime("%d-%m") for d in sorted_days]
    # Asegurar que el día ancla está en day_labels (aunque no tenga movimientos propios)
    if ANCHOR_LBL not in day_labels:
        day_labels.append(ANCHOR_LBL)

    # ANCHOR: último día con ejercicio real per SPEC
    info_29_anchor = days_real[ANCHOR_LBL]["info"]
    last_lbl_tmp = day_labels[-1]
    stock_actual = {sp: info_29_anchor[sp]["u_tot"] for sp in info_29_anchor}

    # Stocks por dia: {lbl: {spec: u_tot}}
    stocks = {last_lbl_tmp: dict(stock_actual)}

    # Si el anchor no coincide con el último día con movimientos, hacer puente:
    # stocks del último día con mov = stocks del anchor (sin movimientos entre ellos)
    if sorted_days:
        last_mov_lbl = sorted_days[-1].strftime("%d-%m")
        if last_mov_lbl != ANCHOR_LBL and last_mov_lbl not in stocks:
            stocks[last_mov_lbl] = dict(stock_actual)

    # Backward propagation: stock[dia anterior] = stock[dia] - entradas[dia] + salidas[dia]
    for d in reversed(sorted_days):
        lbl_after = d.strftime("%d-%m")
        idx = sorted_days.index(d)
        lbl_before = sorted_days[idx-1].strftime("%d-%m") if idx > 0 else "30-04"
        compras = mov_by_day[d]["compras"]
        salidas = mov_by_day[d]["salidas"]
        new_stock = dict(stocks[lbl_after])
        for sp, q in compras.items():
            new_stock[sp] = new_stock.get(sp, 0) - q
        for sp, q in salidas.items():
            new_stock[sp] = new_stock.get(sp, 0) + q
        stocks[lbl_before] = new_stock

    # === FASE 3: dias con inventario REAL usan el rollup real del dia ===
    # La retropropagacion solo puede seguir compras/entregas; no ve el trafico
    # con las OFs (emisiones, recepciones, facturas directas). Como el build ya
    # procesa el inventario completo de CADA ejercicio (days_real), usamos esas
    # cantidades reales por SPEC y dejamos la retropropagacion solo para los
    # dias sin ejercicio (huecos tipo 04..26-may).
    for _lbl_real, _d_real in days_real.items():
        if _lbl_real in stocks or _lbl_real in day_labels:
            stocks[_lbl_real] = {sp: i["u_tot"] for sp, i in _d_real["info"].items()}
    stocks["30-04"] = {sp: i["u_tot"] for sp, i in info_30abr.items()}

    # Detalle de entradas/salidas por SPEC raw y dia (para drill-down al hacer click)
    # entradas_detail[spec][day] = [{proveedor, qty, val, doc, from_spec, qty_per_unit}]
    # salidas_detail[spec][day] = [{cliente, wip, qty_wip, qty_per_unit, qty_raw, val_raw, doc, fuente}]
    entradas_detail = defaultdict(lambda: defaultdict(list))
    salidas_detail = defaultdict(lambda: defaultdict(list))

    for e in entradas_full:
        d_lbl = datetime.fromisoformat(e["fecha"]).strftime("%d-%m")
        canon = e["spec_canon"]
        cu = e["coste_unit"] or 0
        if canon in master.ESCANDALLOS:
            for raw_canon, q in expand_bom_ui(canon).items():
                raw_base = master.base_spec(raw_canon)
                qty_raw = q * e["qty"]
                entradas_detail[raw_base][d_lbl].append({
                    "from_spec": e["spec_emitido"],
                    "proveedor": e["proveedor"],
                    "qty_per_unit": float(q),
                    "qty": qty_raw,
                    "val": qty_raw * cu,
                    "doc": e["doc"],
                })
        else:
            entradas_detail[canon][d_lbl].append({
                "from_spec": e["spec_emitido"],
                "proveedor": e["proveedor"],
                "qty_per_unit": 1.0,
                "qty": e["qty"],
                "val": e["valor"],
                "doc": e["doc"],
            })

    for s in salidas_full:
        d_lbl = datetime.fromisoformat(s["fecha"]).strftime("%d-%m")
        canon = s["spec_canon"]
        cu = s["coste_unit"] or 0
        if canon in master.ESCANDALLOS:
            for raw_canon, q in expand_bom_ui(canon).items():
                raw_base = master.base_spec(raw_canon)
                qty_raw = q * s["qty"]
                salidas_detail[raw_base][d_lbl].append({
                    "wip": s["spec_emitido"],
                    "cliente": s["cliente"],
                    "qty_wip": s["qty"],
                    "qty_per_unit": float(q),
                    "qty_raw": qty_raw,
                    "val_raw": qty_raw * cu,
                    "doc": s["doc"],
                    "fuente": s["fuente"],
                })
        else:
            salidas_detail[canon][d_lbl].append({
                "wip": s["spec_emitido"],
                "cliente": s["cliente"],
                "qty_wip": s["qty"],
                "qty_per_unit": 1.0,
                "qty_raw": s["qty"],
                "val_raw": s["qty"] * cu,
                "doc": s["doc"],
                "fuente": s["fuente"],
            })

    # Calcular gap por SPEC: inferred (último día) - real (anchor 01-jun)
    last_lbl = day_labels[-1]
    info_29 = days_real[ANCHOR_LBL]["info"]  # anchor real
    # Para SPEC en stocks pero no en info_29 o viceversa, manejar
    all_specs = set(stocks[last_lbl]) | set(info_29) | set(info_30abr)
    gap_per_spec = {}
    for sp in all_specs:
        inf = stocks[last_lbl].get(sp, 0)
        real = info_29.get(sp, {"u_tot": 0})["u_tot"]
        gap_per_spec[sp] = inf - real  # positive = inferred too high

    total_gap_uds = sum(gap_per_spec.values())
    # Valor del gap: usamos coste/u del 01-jun (o de 30-abr si no esta)
    def cu_of(sp):
        return (info_29.get(sp, {}).get("unit_cost") or info_30abr.get(sp, {}).get("unit_cost") or 0)
    gap_total_val = sum(v * cu_of(sp) for sp, v in gap_per_spec.items())

    # Aplicar correccion en stocks[30-04] restando el gap a cada spec
    # Equivale a decir "el inventario real 30-abr tenia menos uds que lo que tu fichero dice
    # porque hay Bolsa C+D + escandallos incompletos"
    # NO modificamos stocks[30-04] para no falsear los SPECs individuales.
    # En lugar, anadimos una fila virtual "GAP NO EXPLICADO" que es constante en cada dia.

    # gap_val total = real_29may A+B + C + D vs inferred_29may from SPECs
    # Mas limpio: gap = real_total_30abr - sum(SPEC × cu) en 30-abr
    total_real_30abr = total_30abr_real
    total_inferred_30abr = sum(stocks["30-04"].get(sp, 0) * cu_of(sp) for sp in stocks["30-04"])
    gap_30abr_val = total_real_30abr - total_inferred_30abr

    print(f"  Gap no explicado (constante en todos los dias): {gap_30abr_val:,.2f} EUR")
    print(f"    = real 30-abr ({total_real_30abr:,.0f}) - inferred sum(SPEC x cu) ({total_inferred_30abr:,.0f})")

    print("[5/5] Construyendo HTML...")

    # Reunir info por SPEC para columnas fijas
    all_specs_data = {}
    all_sp_set = set(info_30abr) | set(info_29)
    for _lbl in day_labels:
        all_sp_set |= set(stocks[_lbl].keys())
    for sp in all_sp_set:
        info_src = info_29.get(sp) or info_30abr.get(sp) or {}
        all_specs_data[sp] = {
            "spec": sp,
            "desc": info_src.get("desc", ""),
            "sup": info_src.get("supplier", "(SIN PROVEEDOR)"),
            "tipo": info_src.get("tipo", ""),
            "cu": info_src.get("unit_cost", 0),
        }

    # Minimos por SPEC: fuente = minimos.json (editable desde la app v2 via /api/minimos).
    # Fallback legacy: Live Excel (col G, o col F stock-29-may) solo si el json no existe.
    minimos_pre = {}
    _mj = PROJECT_DIR / "minimos.json"
    if _mj.exists():
        try:
            minimos_pre = {str(k): float(v) for k, v in json.loads(_mj.read_text(encoding="utf-8")).items()}
            print(f"  Minimos: {len(minimos_pre)} SPECs desde minimos.json")
        except Exception as _e:
            print(f"  WARN minimos.json ilegible: {_e}")
    if not minimos_pre:
        try:
            _, minimos_pre, _ = read_live_excel(PROJECT_DIR / "Inventario_Leaseir_Live.xlsx")
        except Exception:
            minimos_pre = {}
        if minimos_pre is None: minimos_pre = {}

    # Construir filas con stocks por dia + movimientos del dia
    # Para movimientos, agregar totales por (spec, day) desde mov_by_day
    spec_compras_dia = defaultdict(lambda: defaultdict(float))
    spec_salidas_dia = defaultdict(lambda: defaultdict(float))
    for d in sorted_days:
        d_lbl = d.strftime("%d-%m")
        for sp, q in mov_by_day[d]["compras"].items():
            spec_compras_dia[sp][d_lbl] += q
        for sp, q in mov_by_day[d]["salidas"].items():
            spec_salidas_dia[sp][d_lbl] += q

    # FASE 3: valorar cada dia real al coste unitario DE ESE DIA (no al del
    # ancla). Para dias sin ejercicio (huecos), coste del ancla como antes.
    def _cu_day(sp, lbl, fallback):
        if lbl == "30-04":
            c = info_30abr.get(sp, {}).get("unit_cost", 0)
        elif lbl in days_real:
            c = days_real[lbl]["info"].get(sp, {}).get("unit_cost", 0)
        else:
            c = 0
        return c or fallback

    rows = []
    for sp, base in all_specs_data.items():
        row = dict(base)
        row["minimo"] = minimos_pre.get(sp)
        for lbl in day_labels:
            u = stocks[lbl].get(sp, 0)
            row[f"u_{lbl}"] = u
            row[f"v_{lbl}"] = u * _cu_day(sp, lbl, base["cu"])
            row[f"entr_{lbl}"] = spec_compras_dia[sp].get(lbl, 0)
            row[f"sal_{lbl}"] = spec_salidas_dia[sp].get(lbl, 0)
        # Total mes
        ent_total = sum(spec_compras_dia[sp].values())
        sal_total = sum(spec_salidas_dia[sp].values())
        row["entr_mes"] = ent_total
        row["sal_mes"] = sal_total
        row["net_mes"] = ent_total - sal_total
        row["entr_mes_val"] = ent_total * base["cu"]
        row["sal_mes_val"] = sal_total * base["cu"]
        row["net_mes_val"] = (ent_total - sal_total) * base["cu"]
        rows.append(row)
    rows.sort(key=lambda r: -r["cu"])

    # Filtrar SPECs con todo 0 en todos los dias
    rows = [r for r in rows if any(abs(r.get(f"u_{lbl}", 0)) > 0.01 for lbl in day_labels)]

    # Split high/others
    stock_high = [r for r in rows if r["cu"] > COST_THRESHOLD]
    stock_low = [r for r in rows if r["cu"] <= COST_THRESHOLD]
    # Agregar others
    others_stock = {lbl: 0 for lbl in day_labels}
    others_stock_val = {lbl: 0 for lbl in day_labels}
    others_mes = {"entr": 0.0, "sal": 0.0, "entr_val": 0.0, "sal_val": 0.0}
    for r in stock_low:
        for lbl in day_labels:
            others_stock[lbl] += r.get(f"u_{lbl}", 0)
            others_stock_val[lbl] += r.get(f"v_{lbl}", 0)
        others_mes["entr"] += r.get("entr_mes", 0)
        others_mes["sal"] += r.get("sal_mes", 0)
        others_mes["entr_val"] += r.get("entr_mes_val", 0)
        others_mes["sal_val"] += r.get("sal_mes_val", 0)

    # GAP row: constante en todos los dias, en valor
    # Reales por dia (donde aplique)
    reales_check = {"30-04": total_30abr_real}
    for lbl_, info in days_real.items():
        reales_check[lbl_] = info["total"]

    # Ajuste por dia: real_total - sum(SPECs × cu). Para dias con real cuadra;
    # para intermedios usamos el valor de 01-jun como aproximacion.
    ajuste_anchor = reales_check[ANCHOR_LBL] - (sum(r.get(f"v_{ANCHOR_LBL}", 0) for r in stock_high) + others_stock_val[ANCHOR_LBL])
    ajuste_per_day = {}
    for lbl in day_labels:
        val_specs = sum(r.get(f"v_{lbl}", 0) for r in stock_high) + others_stock_val[lbl]
        if lbl in reales_check:
            ajuste_per_day[lbl] = reales_check[lbl] - val_specs
        else:
            ajuste_per_day[lbl] = ajuste_anchor

    gap_row = {
        "spec": "AJUSTE (C+D)",
        "desc": "Bolsas C+D (overhead, WIPs sin BOM) + escandallos incompletos. Cuadrado contra real en 30-abr, 27/28/29-may y 01-jun. Resto estimado.",
        "sup": "(contable)",
        "tipo": "Ajuste",
        "cu": 0,
    }
    for lbl in day_labels:
        gap_row[f"u_{lbl}"] = 0
        gap_row[f"v_{lbl}"] = ajuste_per_day[lbl]

    # Totales por dia
    totales = {}
    for lbl in day_labels:
        val_specs = sum(r.get(f"v_{lbl}", 0) for r in stock_high) + others_stock_val[lbl]
        totales[lbl] = val_specs + ajuste_per_day[lbl]

    # Reales por dia (donde aplique)
    reales = {"30-04": total_30abr_real}
    for lbl, info in days_real.items():
        reales[lbl] = info["total"]

    # Stocks de WIPs (no aparecen en info porque build_totales decompone vía BOM).
    # Leemos directamente del fichero inventario los WIPs presentes en ESCANDALLOS.
    # Precomputar set de WIP bases una vez
    wip_canons_set = set(master.ESCANDALLOS or {})
    wip_bases_set = set(master.base_spec(k) for k in wip_canons_set)

    def load_wip_stocks(daily_file):
        """Devuelve dict {wip_base: qty} para WIPs en ESCANDALLOS presentes en inventario
        + OFs en vuelo (qty estimada via emitido/cu del WIP)."""
        wb = openpyxl.load_workbook(daily_file, data_only=True)
        wips = {}
        # 1) Inventario directo (stock fisico de WIPs terminados)
        if "inventario" in wb.sheetnames:
            sh = wb["inventario"]
            # Cu por SPEC para estimar qty OFs despues
            cu_by_canon = {}
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 4 or not row[0]: continue
                code = str(row[0]).strip()
                canon = normalize_code(code)
                base = master.base_spec(canon)
                try: qty = float(row[2] or 0)
                except: qty = 0
                try: val = float(row[3] or 0)
                except: val = 0
                if qty > 0 and val > 0:
                    cu_by_canon[canon] = val / qty
                    cu_by_canon.setdefault(base, val / qty)
                if canon in wip_canons_set or base in wip_bases_set:
                    if qty > 0:
                        wips[base] = wips.get(base, 0) + qty
            # 2) OFs en vuelo (WIPs en produccion) — estimar qty via emitido/cu
            if "en vuelo" in wb.sheetnames:
                sh2 = wb["en vuelo"]
                for row in sh2.iter_rows(values_only=True, min_row=2):
                    if not row or len(row) < 8: continue
                    spec_code = str(row[2] or "").strip()
                    if not spec_code: continue
                    canon = normalize_code(spec_code)
                    base = master.base_spec(canon)
                    if canon not in wip_canons_set and base not in wip_bases_set:
                        continue
                    try: emitido = float(row[5] or 0)
                    except: emitido = 0
                    try: recibido = float(row[6] or 0)
                    except: recibido = 0
                    try: pendiente = float(row[7] or 0)
                    except: pendiente = 0
                    if pendiente <= 0: continue
                    cu = cu_by_canon.get(canon) or cu_by_canon.get(base) or 0
                    if cu <= 0:
                        # estimar via coste teorico del escandallo
                        try:
                            exp = expand_bom_ui(canon if canon in master.ESCANDALLOS else base)
                            cu = sum(q * (cu_by_canon.get(r, 0) or cu_by_canon.get(master.base_spec(r), 0)) for r, q in exp.items())
                        except: cu = 0
                    if cu > 0:
                        qty_total = max(1, round(emitido / cu)) if emitido > 0 else 1
                        qty_recib = round(recibido / cu) if recibido > 0 else 0
                        qty_pend = max(1, qty_total - qty_recib)
                    else:
                        qty_pend = 1
                    wips[base] = wips.get(base, 0) + qty_pend
        return wips

    # WIP stocks: solo cargar el día ancla (último) para minimizar tiempo; resto reutiliza fallback
    wip_stocks_per_day = {}
    # Cargar wip_stocks de TODOS los ejercicios disponibles para que el breakdown raw/embebido
    # del drill-down funcione correctamente cualquier día seleccionado.
    for _iso_w, _lbl_w, _fn_w in ej_files:
        try:
            wip_stocks_per_day[_lbl_w] = load_wip_stocks(_fn_w)
        except Exception as _e:
            print(f"  WARN wip_stocks {_lbl_w}: {_e}")

    # Inverse BOM: raw_canon → [{wip: spec_canon, qty: qty_per_unit, wip_desc: ...}]
    # Permite calcular embed dinámico en JS: para cualquier día, embed_raw = sum(u_wip × qty)
    # IMPORTANTE: usar base_spec(wip) y agregar por base para que coincida con keys de stock
    inv_bom_tmp = defaultdict(lambda: defaultdict(float))  # raw_base → {wip_base: qty_acumulada}
    for wip_canon in (master.ESCANDALLOS or {}).keys():
        try:
            exp = expand_bom_ui(wip_canon)
        except Exception:
            continue
        wip_base = master.base_spec(wip_canon)
        for raw_canon, q in exp.items():
            raw_base = master.base_spec(raw_canon)
            inv_bom_tmp[raw_base][wip_base] += float(q)
    # Ahora convertir a lista de dicts incluyendo desc — SOLO WIPs con stock real en algún anchor
    wip_with_stock = set()
    for wips_day in wip_stocks_per_day.values():
        wip_with_stock.update(wips_day.keys())
    inv_bom = {}
    for raw_base, wip_qty in inv_bom_tmp.items():
        rows_list = []
        for wip_base, q in wip_qty.items():
            if wip_base not in wip_with_stock:
                continue  # saltar WIPs sin stock (no aportan al embed)
            wip_info = all_specs_data.get(wip_base, {})
            rows_list.append({
                "wip": wip_base,
                "qty": q,
                "wip_desc": (wip_info.get("desc", "") or "")[:40],
            })
        if not rows_list:
            continue
        rows_list.sort(key=lambda x: -x["qty"])
        inv_bom[raw_base] = rows_list

    # Chart series por SPEC
    chart_series = {}
    for r in stock_high:
        sp = r["spec"]
        chart_series[sp] = {
            "labels": day_labels,
            "u_tot": [r.get(f"u_{lbl}", 0) for lbl in day_labels],
            "val": [r.get(f"v_{lbl}", 0) for lbl in day_labels],
            "desc": r["desc"], "sup": r["sup"], "tipo": r["tipo"], "cu": r["cu"],
            "minimo": r.get("minimo"),
        }

    # Construir entradas/salidas para tabs (con desglose diario)
    # day_labels[0] = "30-04" (no hay mov), resto son dias con mov
    mov_day_labels = day_labels[1:]
    compras_agg = defaultdict(float)
    salidas_agg = defaultdict(float)
    compras_by_day = defaultdict(lambda: defaultdict(float))  # spec -> day -> qty
    salidas_by_day = defaultdict(lambda: defaultdict(float))
    for d in sorted_days:
        lbl = d.strftime("%d-%m")
        for sp, q in mov_by_day[d]["compras"].items():
            compras_agg[sp] += q
            compras_by_day[sp][lbl] += q
        for sp, q in mov_by_day[d]["salidas"].items():
            salidas_agg[sp] += q
            salidas_by_day[sp][lbl] += q

    # Totales agregados por dia para footer
    compras_totales_dia = defaultdict(float)
    salidas_totales_dia = defaultdict(float)
    for sp in compras_agg:
        for lbl in mov_day_labels:
            compras_totales_dia[lbl] += compras_by_day[sp].get(lbl, 0)
    for sp in salidas_agg:
        for lbl in mov_day_labels:
            salidas_totales_dia[lbl] += salidas_by_day[sp].get(lbl, 0)

    entradas_rows = []
    entradas_rows_low = []
    others_entradas = {"u_tot": 0, "val": 0}
    others_entradas_dia = defaultdict(float)
    others_entradas_dia_val = defaultdict(float)
    for sp, q in compras_agg.items():
        info = all_specs_data.get(sp, {})
        cu = info.get("cu", 0)
        val = cu * q
        row = {"spec": sp, "desc": info.get("desc", ""), "sup": info.get("sup", ""),
               "tipo": info.get("tipo", ""), "cu": cu, "uds": q, "val": val,
               "u_stock": stocks[last_lbl].get(sp, 0), "v_stock": stocks[last_lbl].get(sp, 0) * cu}
        for lbl in mov_day_labels:
            row[f"u_{lbl}"] = compras_by_day[sp].get(lbl, 0)
            row[f"v_{lbl}"] = compras_by_day[sp].get(lbl, 0) * cu
        if cu > COST_THRESHOLD:
            entradas_rows.append(row)
        else:
            entradas_rows_low.append(row)
            others_entradas["u_tot"] += q
            others_entradas["val"] += val
            for lbl in mov_day_labels:
                q_day = compras_by_day[sp].get(lbl, 0)
                others_entradas_dia[lbl] += q_day
                others_entradas_dia_val[lbl] += q_day * cu
    entradas_rows.sort(key=lambda r: -r["val"])
    entradas_rows_low.sort(key=lambda r: -r["val"])

    salidas_rows = []
    salidas_rows_low = []
    others_salidas_t = {"u_tot": 0, "val": 0}
    others_salidas_dia = defaultdict(float)
    others_salidas_dia_val = defaultdict(float)
    for sp, q in salidas_agg.items():
        info = all_specs_data.get(sp, {})
        cu = info.get("cu", 0)
        val = cu * q
        row = {"spec": sp, "desc": info.get("desc", ""), "sup": info.get("sup", ""),
               "tipo": info.get("tipo", ""), "cu": cu, "uds": q, "val": val,
               "u_stock": stocks[last_lbl].get(sp, 0), "v_stock": stocks[last_lbl].get(sp, 0) * cu}
        for lbl in mov_day_labels:
            row[f"u_{lbl}"] = salidas_by_day[sp].get(lbl, 0)
            row[f"v_{lbl}"] = salidas_by_day[sp].get(lbl, 0) * cu
        if cu > COST_THRESHOLD:
            salidas_rows.append(row)
        else:
            salidas_rows_low.append(row)
            others_salidas_t["u_tot"] += q
            others_salidas_t["val"] += val
            for lbl in mov_day_labels:
                q_day = salidas_by_day[sp].get(lbl, 0)
                others_salidas_dia[lbl] += q_day
                others_salidas_dia_val[lbl] += q_day * cu
    salidas_rows.sort(key=lambda r: -r["val"])
    salidas_rows_low.sort(key=lambda r: -r["val"])

    # info_29_local apunta al anchor real (usado por todas las cliente tables)
    info_29_local = days_real[ANCHOR_LBL]["info"]
    # Fallback global de proveedores: hoja "proveedores" del ejercicio anchor.
    # Cubre TODOS los SPECs (incluidos WIPs vendidos) que no estan en stock.
    suppliers_anchor = {}
    try:
        _wb_sup = openpyxl.load_workbook(ej_files[-1][2], data_only=True, read_only=True)
        _sh_sup_name = next((n for n in _wb_sup.sheetnames if n.lower().startswith("proveedor")), None)
        if _sh_sup_name:
            _sh_sup = _wb_sup[_sh_sup_name]
            _by_code = defaultdict(set)
            for _r in _sh_sup.iter_rows(values_only=True, min_row=2):
                if not _r or not _r[0]:
                    continue
                _code = str(_r[0]).strip()
                _sup = str((_r[3] if len(_r) > 3 else "") or "").strip()
                if not _sup:
                    continue
                _by_code[normalize_code(_code)].add(_sup)
                _by_code[_code].add(_sup)
            for _k, _names in _by_code.items():
                suppliers_anchor[_k] = next(iter(_names)) if len(_names) == 1 else " / ".join(sorted(_names))
        _wb_sup.close()
    except Exception as _e:
        print(f"  WARN: no se pudo cargar hoja proveedores del anchor: {_e}")

    def _lookup_supplier(spec_emit):
        canon = normalize_code(spec_emit)
        base = master.base_spec(canon)
        # 1) rollup del dia (raws en stock)
        info_sp = info_29_local.get(canon) or info_29_local.get(base) or {}
        s = info_sp.get("supplier", "") or ""
        if s and s != "(SIN PROVEEDOR)":
            return s
        # 2) hoja "proveedores" del ejercicio (lookup directo)
        s2 = suppliers_anchor.get(canon) or suppliers_anchor.get(base)
        if s2:
            return s2
        # 3) si es WIP, agregamos el proveedor mas usado entre sus raws (ponderado por qty)
        if canon in master.ESCANDALLOS:
            try:
                exp = expand_bom_ui(canon)
            except Exception:
                exp = {}
            from collections import Counter as _Cnt
            cnt = _Cnt()
            for _raw, _q in exp.items():
                _info_r = info_29_local.get(_raw) or info_29_local.get(master.base_spec(_raw)) or {}
                _sup_r = _info_r.get("supplier", "") or ""
                if not _sup_r or _sup_r == "(SIN PROVEEDOR)":
                    _sup_r = suppliers_anchor.get(_raw) or suppliers_anchor.get(master.base_spec(_raw)) or ""
                if _sup_r and _sup_r != "(SIN PROVEEDOR)":
                    cnt[_sup_r] += _q
            if cnt:
                return cnt.most_common(1)[0][0]
        return s or ""

    # Salidas por cliente (WIP-level, agg) — incluye qty/val por día
    agg_cli = defaultdict(lambda: {"qty": 0, "val": 0, "coste_unit": 0, "desc": "", "lotes": set(),
                                    "by_day_q": defaultdict(float), "by_day_v": defaultdict(float)})
    for s in salidas_full:
        k = (s["cliente"], s["spec_emitido"])
        d_lbl = datetime.fromisoformat(s["fecha"]).strftime("%d-%m")
        agg_cli[k]["qty"] += s["qty"]
        agg_cli[k]["val"] += s["valor"]
        agg_cli[k]["coste_unit"] = s["coste_unit"]
        agg_cli[k]["desc"] = s["desc"]
        if s["lote"]: agg_cli[k]["lotes"].add(s["lote"])
        agg_cli[k]["by_day_q"][d_lbl] += s["qty"]
        agg_cli[k]["by_day_v"][d_lbl] += s["valor"]
    salidas_cliente_table = []
    for (cli, sp), v in agg_cli.items():
        sup = _lookup_supplier(sp)
        row = {
            "cliente": cli, "spec": sp, "desc": v["desc"], "sup": sup,
            "qty": v["qty"], "coste_unit": v["coste_unit"], "val": v["val"],
            "lotes": ", ".join(sorted(v["lotes"]))[:60],
        }
        for d_lbl, q in v["by_day_q"].items():
            row[f"qty_{d_lbl}"] = q
            row[f"val_{d_lbl}"] = v["by_day_v"][d_lbl]
        salidas_cliente_table.append(row)
    salidas_cliente_table.sort(key=lambda r: -r["val"])

    # Salidas por cliente vista RAW: descomponer WIPs a raws (con qty/val por día)
    agg_cli_raw = defaultdict(lambda: {"qty": 0, "val": 0, "via": defaultdict(float),
                                        "by_day_q": defaultdict(float), "by_day_v": defaultdict(float)})
    for s in salidas_full:
        canon = s["spec_canon"]
        qty_wip = s["qty"]
        cli = s["cliente"]
        d_lbl = datetime.fromisoformat(s["fecha"]).strftime("%d-%m")
        if canon in master.ESCANDALLOS:
            try:
                exp = expand_bom_ui(canon)
            except Exception:
                exp = None
            if exp:
                for raw_canon, q in exp.items():
                    raw_base = master.base_spec(raw_canon)
                    qty_raw = q * qty_wip
                    cu_raw = info_29_local.get(raw_base, {}).get("unit_cost", 0)
                    k = (cli, raw_base)
                    agg_cli_raw[k]["qty"] += qty_raw
                    agg_cli_raw[k]["val"] += qty_raw * cu_raw
                    agg_cli_raw[k]["via"][s["spec_emitido"]] += qty_wip
                    agg_cli_raw[k]["by_day_q"][d_lbl] += qty_raw
                    agg_cli_raw[k]["by_day_v"][d_lbl] += qty_raw * cu_raw
                continue
        # Fallback: raw directo (sin BOM)
        cu_raw = info_29_local.get(canon, {}).get("unit_cost", s["coste_unit"] or 0)
        k = (cli, canon)
        agg_cli_raw[k]["qty"] += qty_wip
        agg_cli_raw[k]["val"] += qty_wip * cu_raw
        agg_cli_raw[k]["via"][s["spec_emitido"]] += qty_wip
        agg_cli_raw[k]["by_day_q"][d_lbl] += qty_wip
        agg_cli_raw[k]["by_day_v"][d_lbl] += qty_wip * cu_raw
    salidas_cliente_raw_table = []
    for (cli, raw_sp), v in agg_cli_raw.items():
        info_sp = info_29_local.get(raw_sp, {})
        row = {
            "cliente": cli, "spec": raw_sp,
            "desc": info_sp.get("desc", "")[:80],
            "sup": info_sp.get("supplier", ""),
            "tipo": info_sp.get("tipo", ""),
            "qty": v["qty"], "coste_unit": info_sp.get("unit_cost", 0),
            "val": v["val"],
            "via": ", ".join(f"{k}({int(round(v_))})" for k, v_ in sorted(v["via"].items(), key=lambda x: -x[1])[:3]),
        }
        for d_lbl, q in v["by_day_q"].items():
            row[f"qty_{d_lbl}"] = q
            row[f"val_{d_lbl}"] = v["by_day_v"][d_lbl]
        salidas_cliente_raw_table.append(row)
    salidas_cliente_raw_table.sort(key=lambda r: -r["val"])

    # Salidas por WIP (vista WIP de Salidas mes): aggregar por spec_emitido
    agg_wip = defaultdict(lambda: {"qty": 0, "val": 0, "desc": "", "clientes": set()})
    # Necesitamos qty/val por día para poder filtrar por mes en el front
    agg_wip_by_day = defaultdict(lambda: defaultdict(lambda: {"qty": 0.0, "val": 0.0}))
    for s in salidas_full:
        sp = s["spec_emitido"]
        d_lbl = datetime.fromisoformat(s["fecha"]).strftime("%d-%m")
        agg_wip[sp]["qty"] += s["qty"]
        agg_wip[sp]["val"] += s["valor"]
        agg_wip[sp]["desc"] = s["desc"]
        agg_wip[sp]["clientes"].add(s["cliente"])
        agg_wip_by_day[sp][d_lbl]["qty"] += s["qty"]
        agg_wip_by_day[sp][d_lbl]["val"] += s["valor"]
    salidas_wip_table = []
    for sp, v in agg_wip.items():
        canon = normalize_code(sp)
        info_sp = info_29_local.get(master.base_spec(canon), {}) or info_29_local.get(canon, {})
        row = {
            "spec": sp, "desc": v["desc"],
            "sup": info_sp.get("supplier", ""),
            "tipo": info_sp.get("tipo", ""),
            "cu": v["val"] / v["qty"] if v["qty"] else 0,
            "uds": v["qty"], "val": v["val"],
            "u_stock": info_sp.get("u_tot", 0),
            "v_stock": info_sp.get("u_tot", 0) * info_sp.get("unit_cost", 0),
            "clientes": ", ".join(sorted(v["clientes"]))[:60],
        }
        # Por día (formato sal_DD-MM / val_DD-MM)
        for d_lbl, agg in agg_wip_by_day[sp].items():
            row[f"sal_{d_lbl}"] = agg["qty"]
            row[f"valsal_{d_lbl}"] = agg["val"]
        salidas_wip_table.append(row)
    salidas_wip_table.sort(key=lambda r: -r["val"])

    # ---- SIMULATION ----
    LIVE_FILE = PROJECT_DIR / "Inventario_Leaseir_Live.xlsx"
    forecast_sim, minimos_sim, compras_obl_sim = read_live_excel(LIVE_FILE)
    sim_data = {}
    if forecast_sim is not None:
        stock_29_dict = {}
        for sp in info_29:
            info = info_29[sp]
            stock_29_dict[sp] = {
                "u_tot": info.get("u_tot", 0),
                "cu": info.get("unit_cost", 0),
                "desc": info.get("desc", ""),
                "supplier": info.get("supplier", ""),
                "tipo": info.get("tipo", ""),
            }
        sim_data = compute_simulation(stock_29_dict, forecast_sim, minimos_sim, compras_obl_sim)
        print(f"  Simulación calculada: {len(sim_data)} SPECs con movimiento")
    else:
        print("  Inventario_Leaseir_Live.xlsx no encontrado, simulación vacía")

    # raw_real_per_day[day][spec] = u_raw (bolsa A sueltos en hoja inventario)
    raw_real_per_day = {}
    for _lbl_r, _d_r in days_real.items():
        _per_sp = {}
        for _sp_r, _i_r in _d_r["info"].items():
            _ur = _i_r.get("u_raw")
            if _ur is not None:
                _per_sp[_sp_r] = _ur
        raw_real_per_day[_lbl_r] = _per_sp

    # === FASE 3: flujo de fabrica (OFs en vuelo), diff dia a dia ===
    # Por cada OF: Δemitido (almacen+facturas -> OF) y Δrecibido (OF -> almacen)
    # entre ejercicios consecutivos. OF que desaparece = cerrada (su pendiente
    # restante se apunta como recibido/cerrado ese dia).
    of_days = []
    of_snap = {}
    for _iso_o, _lbl_o, _fn_o in ej_files:
        of_snap[_lbl_o] = load_ofs_en_vuelo_detalle(_fn_o)
        of_days.append(_lbl_o)
    of_rows_agg = {}
    of_pend_tot = {}
    _prev_o = None
    for _lbl_o in of_days:
        snap = of_snap[_lbl_o]
        of_pend_tot[_lbl_o] = sum(o["pendiente"] for o in snap.values())
        for of_num, o in snap.items():
            r = of_rows_agg.setdefault(of_num, {
                "of": of_num, "spec": o["spec"], "desc": o["desc"], "status": o["status"],
                "pend": {}, "de": {}, "dr": {}, "emitido": 0.0, "recibido": 0.0,
                "last_seen": _lbl_o, "closed": False})
            r["pend"][_lbl_o] = o["pendiente"]
            r["emitido"] = o["emitido"]
            r["recibido"] = o["recibido"]
            r["status"] = o["status"] or r["status"]
            r["last_seen"] = _lbl_o
            r["closed"] = False
            if _prev_o is not None:
                po = of_snap[_prev_o].get(of_num)
                de = o["emitido"] - (po["emitido"] if po else 0.0)
                dr = o["recibido"] - (po["recibido"] if po else 0.0)
                if abs(de) > 0.005:
                    r["de"][_lbl_o] = de
                if abs(dr) > 0.005:
                    r["dr"][_lbl_o] = dr
        if _prev_o is not None:
            for of_num, po in of_snap[_prev_o].items():
                if of_num not in snap and of_num in of_rows_agg and not of_rows_agg[of_num]["closed"]:
                    rc = of_rows_agg[of_num]
                    rc["closed"] = True
                    if po["pendiente"] > 0.005:
                        rc["dr"][_lbl_o] = rc["dr"].get(_lbl_o, 0.0) + po["pendiente"]
                    rc["pend"][_lbl_o] = 0.0
        _prev_o = _lbl_o
    of_rows_list = sorted(of_rows_agg.values(),
                          key=lambda r: -(r["pend"].get(r["last_seen"], 0) + sum(abs(v) for v in r["de"].values())))
    print(f"  Fase 3: {len(of_rows_list)} OFs seguidas en 'en vuelo'; pendiente ancla = {of_pend_tot.get(of_days[-1] if of_days else '', 0):,.0f} EUR")

    # === FASE 3.1: vigilancia de PRECIOS DE FICHA ===
    # Para cada WIP fisico del inventario del ancla: precio de ficha SAP
    # (valor/cantidad en libros) vs suma real de sus materiales al coste de HOY
    # (precios de compra reales via PURCHASE_HISTORY). Se recalcula en cada
    # corrida: si los costes de compra cambian, la desviacion se reajusta sola.
    # Sin mano de obra, overhead ni servicios externos (confirmado), toda
    # desviacion estable aqui es una ficha a revisar en SAP.
    def _cu_anchor_f(sp):
        _i = info_29.get(sp) or info_29.get(master.base_spec(sp)) or {}
        return _i.get("unit_cost", 0) or 0

    def _materiales_f(canon):
        if canon in master.ESCANDALLOS:
            return sum(q * _cu_anchor_f(r) for r, q in expand_bom_ui(canon).items())
        rec_q = getattr(master, "QUARANTINED_RECIPES", {}).get(canon)
        if rec_q is not None:
            tot = 0.0
            for child, q in rec_q:
                if child in master.ESCANDALLOS:
                    tot += q * sum(qq * _cu_anchor_f(r) for r, qq in expand_bom_ui(child).items())
                else:
                    tot += q * _cu_anchor_f(child)
            return tot
        return None

    fichas_rows = []
    try:
        _wb_f = openpyxl.load_workbook(ej_files[-1][2], data_only=True, read_only=True)
        _sh_f = next((n for n in _wb_f.sheetnames if n.lower().strip() == "inventario"), None)
        _agg_f = {}
        if _sh_f:
            for _r in _wb_f[_sh_f].iter_rows(values_only=True, min_row=2):
                if not _r or not _r[0] or len(_r) < 4:
                    continue
                _code = str(_r[0]).strip()
                _canon = normalize_code(_code)
                _base = master.base_spec(_canon)
                _quar = getattr(master, "QUARANTINED_RECIPES", {})
                if _canon not in master.ESCANDALLOS and _canon not in _quar and _base not in master.QUARANTINED_BOMS:
                    continue
                try:
                    _q = float(_r[2] or 0)
                    _v = float(_r[3] or 0)
                except (TypeError, ValueError):
                    continue
                if _q <= 0:
                    continue
                _a = _agg_f.setdefault(_base, {"qty": 0.0, "val": 0.0, "canon": _canon,
                                               "desc": str(_r[1] or "")[:60]})
                _a["qty"] += _q
                _a["val"] += _v
        _wb_f.close()
        for _b, _a in _agg_f.items():
            _ficha = _a["val"] / _a["qty"]
            _mat = _materiales_f(_a["canon"])
            if _mat is None or _mat <= 0:
                continue
            _dif = _ficha - _mat
            if abs(_dif) < 5 or abs(_dif * _a["qty"]) < 150:
                continue
            # Parte EXPLICADA de la diferencia = su PATRON NORMAL historico
            # (FICHAS_NORMAL, mediana de 50 ejercicios). Fallback: contenido
            # de lineas PROC de su expansion. El residual = dif - patron es la
            # unica cifra a vigilar (solo crece si algo se desvia de nuevo).
            _refurb = _b in master.QUARANTINED_BOMS
            _normal = getattr(master, "FICHAS_NORMAL", {}).get(_b)
            if _normal is None:
                if _refurb:
                    _normal = _dif
                else:
                    _normal = 0.0
                    if _a["canon"] in master.ESCANDALLOS:
                        for _rr, _qq in master.expand_bom(_a["canon"]).items():
                            if str(_rr).startswith("PROC-"):
                                _normal += _qq * master.PROCESS_COSTS.get(str(_rr)[5:], 0.0)
            _expl = _normal
            _resid = _dif - _expl
            fichas_rows.append({"spec": _b, "desc": _a["desc"], "qty": _a["qty"],
                                "ficha": _ficha, "mat": _mat, "dif": _dif,
                                "expl": _expl, "resid": _resid,
                                "impacto": _dif * _a["qty"],
                                "imp_resid": _resid * _a["qty"],
                                "cuarentena": _refurb})
        fichas_rows.sort(key=lambda r: -abs(r["imp_resid"]))
    except Exception as _e:
        print(f"  WARN fichas: {_e}")
    fichas_total = sum(r["impacto"] for r in fichas_rows)
    fichas_resid = sum(r["imp_resid"] for r in fichas_rows)
    print(f"  Recetas vs libros: {len(fichas_rows)} filas; dif bruta neta = {fichas_total:,.0f} EUR; RESIDUAL no explicado = {fichas_resid:,.0f} EUR")

    # BOM directo (v2/Prevision): parent_base -> {"desc", "mat": [[raw_base, qty], ...]}
    # Expansion completa sin lineas PROC-* (unidades fisicas).
    _desc_lookup = {}
    for _r in rows:
        if _r.get("desc"):
            _desc_lookup.setdefault(_r["spec"], _r["desc"])
    bom_fwd = {}
    for _parent in sorted(master.ESCANDALLOS.keys()):
        _pb = master.base_spec(_parent)
        if _pb in bom_fwd:
            continue
        try:
            _exp = expand_bom_ui(_parent)
        except Exception:
            continue
        _mats_agg = defaultdict(float)
        for _rc, _q in _exp.items():
            _mats_agg[master.base_spec(_rc)] += float(_q)
        if _mats_agg:
            bom_fwd[_pb] = {
                "desc": _desc_lookup.get(_pb, ""),
                "mat": [[_rb, round(_q, 4)] for _rb, _q in sorted(_mats_agg.items())],
            }
    print(f"  BOM directo v2: {len(bom_fwd)} piezas con receta")

    payload = {
        "days": day_labels, "last": last_lbl,
        "real_days": ["30-04"] + of_days,
        "of_flows": {"days": of_days, "rows": of_rows_list},
        "of_pend": of_pend_tot,
        "fichas": {"rows": fichas_rows, "total": fichas_total, "resid": fichas_resid,
                   "dia": ej_files[-1][1] if ej_files else ""},
        "mov_days": mov_day_labels,
        "stock": stock_high,
        "others_stock_u": others_stock,
        "others_stock_v": others_stock_val,
        "others_mes": others_mes,
        "gap_row": gap_row,
        "totales_inferred": totales,
        "totales_real": reales,
        "entradas": entradas_rows, "others_entradas": others_entradas,
        "entradas_low": entradas_rows_low,
        "others_entradas_dia": dict(others_entradas_dia),
        "others_entradas_dia_val": dict(others_entradas_dia_val),
        "salidas": salidas_rows, "others_salidas": others_salidas_t,
        "salidas_low": salidas_rows_low,
        "others_salidas_dia": dict(others_salidas_dia),
        "others_salidas_dia_val": dict(others_salidas_dia_val),
        "stock_low": stock_low,
        "compras_totales_dia": dict(compras_totales_dia),
        "salidas_totales_dia": dict(salidas_totales_dia),
        "salidas_cliente": salidas_cliente_table,
        "salidas_cliente_raw": salidas_cliente_raw_table,
        "salidas_wip": salidas_wip_table,
        "chart_series": chart_series,
        "entradas_detail": {k: dict(v) for k, v in entradas_detail.items()},
        "salidas_detail": {k: dict(v) for k, v in salidas_detail.items()},
        "cost_threshold": COST_THRESHOLD,
        "gap_30abr_val": gap_30abr_val,
        "simulation": sim_data,
        "inv_bom": inv_bom,
        "wip_stocks": wip_stocks_per_day,
        "raw_real_per_day": raw_real_per_day,
        "bom": bom_fwd,
    }

    import time as _time
    _t = _time.time()
    data_json = json.dumps(payload, ensure_ascii=False, default=float)
    print(f"  json.dumps: {_time.time()-_t:.1f}s, size={len(data_json)/1024/1024:.1f}MB")

    _t = _time.time()
    template = TEMPLATE.read_text(encoding="utf-8")
    out = template.replace("__DATA__", data_json)
    print(f"  template replace: {_time.time()-_t:.1f}s")

    import shutil
    _t = _time.time()
    tmp_path = Path("/tmp") / OUTPUT_HTML.name
    tmp_path.write_text(out, encoding="utf-8")
    print(f"  tmp write: {_time.time()-_t:.1f}s")

    _t = _time.time()
    shutil.copyfile(tmp_path, OUTPUT_HTML)
    print(f"  copy to OneDrive: {_time.time()-_t:.1f}s")

    # v2 (Vercel) — mismo payload, template extendido con Alertas + Prevision
    if TEMPLATE_V2.exists():
        _t = _time.time()
        out2 = TEMPLATE_V2.read_text(encoding="utf-8").replace("__DATA__", data_json)
        tmp2 = Path("/tmp") / OUTPUT_HTML_V2.name
        tmp2.write_text(out2, encoding="utf-8")
        shutil.copyfile(tmp2, OUTPUT_HTML_V2)
        print(f"  v2 render: {_time.time()-_t:.1f}s -> {OUTPUT_HTML_V2.name}")

    print(f"  Entradas: {len(entradas_rows)} + {len(entradas_rows_low)} Others")
    print(f"  Salidas: {len(salidas_rows)} + {len(salidas_rows_low)} Others")


if __name__ == "__main__":
    main()
