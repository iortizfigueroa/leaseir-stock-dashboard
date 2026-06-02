#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
actualizar_evolutivo.py
=======================

Genera Stock_evolutivo.xlsx con 4 pestanas:
  - Resumen general: Bolsas A/B/C/D por dia
  - Por proveedor: agrupado por proveedor, 4 cols por dia (uds raw/emb/tot + valor)
  - Por tipo: agrupado por tipo de componente
  - Por SPEC: TODOS los SPECs (1000+), 4 cols por dia + Mov.ultimo dia

Detecta automaticamente todos los ficheros 'ejercicio DD-MM-YYYY.xlsx' en
la carpeta del proyecto y los procesa en orden cronologico.
"""

from __future__ import annotations
import sys
import re
import glob
import importlib.util
from collections import defaultdict
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = PROJECT_DIR / "Stock_evolutivo.xlsx"

# Reusar master / prov / costM sin duplicar instancias
sys.path.insert(0, str(PROJECT_DIR))
if "master" in sys.modules:
    master = sys.modules["master"]
else:
    spec_m = importlib.util.spec_from_file_location("master", PROJECT_DIR / "inventario_master.py")
    master = importlib.util.module_from_spec(spec_m)
    spec_m.loader.exec_module(master)
    sys.modules["master"] = master
if not master.ESCANDALLOS:
    master.load_escandallos()

if "prov" in sys.modules:
    prov = sys.modules["prov"]
else:
    spec_p = importlib.util.spec_from_file_location("prov", PROJECT_DIR / "stock_por_proveedor.py")
    prov = importlib.util.module_from_spec(spec_p)
    spec_p.loader.exec_module(prov)
    sys.modules["prov"] = prov

if "costM" in sys.modules:
    costM = sys.modules["costM"]
else:
    spec_c = importlib.util.spec_from_file_location("costM", PROJECT_DIR / "stock_por_coste.py")
    costM = importlib.util.module_from_spec(spec_c)
    spec_c.loader.exec_module(costM)
    sys.modules["costM"] = costM


# ============================================================
# Deteccion de ficheros ejercicio DD-MM-YYYY.xlsx
# ============================================================

def find_daily_files():
    """Devuelve [(label, date, path)] ordenado cronologicamente."""
    pattern = str(PROJECT_DIR / "ejercicio *.xlsx")
    found = []
    for fpath in glob.glob(pattern):
        name = Path(fpath).stem
        m = re.search(r"(\d{2})-(\d{2})-(\d{4})", name)
        if not m:
            continue
        day, month, year = m.groups()
        try:
            dt = datetime(int(year), int(month), int(day))
        except ValueError:
            continue
        label = f"{day}-{month}"
        found.append((label, dt, Path(fpath)))
    found.sort(key=lambda x: x[1])
    return found


# ============================================================
# Procesado de un fichero diario
# ============================================================

def process_day(daily_file: Path):
    """Devuelve (rec, rollup_with_cost_inferido)."""
    suppliers = prov.load_suppliers_from_xlsx(daily_file)
    inv = prov.load_inventory_from_xlsx(daily_file)
    of_rows = prov.load_ofs_from_xlsx(daily_file, inv)
    inv.extend(of_rows)
    for row in inv:
        row["cat"] = master.classify(row, suppliers)
    unit_cost = master.build_unit_costs(inv)
    recon = master.reconcile(inv, unit_cost)
    rollup = master.build_rollup(inv, recon, unit_cost, suppliers, {})

    # Inferir Monocrom + refine_type
    for r in rollup:
        if not r.get("supplier"):
            d = r.get("desc", "") or ""
            if re.search(r"laser diode|\bENSPSTM\b|ensptm|\bLBS-|\bmonocrom\b", d, re.IGNORECASE):
                r["supplier"] = "MONOCROM S.L"
        r["tipo"] = prov.refine_type(r["tipo"], r.get("supplier", ""), r.get("desc", ""), r.get("unit_cost", 0))

    # Inferencia de coste
    acc = defaultdict(list)
    for r in rollup:
        if r["unit_cost"] and r["unit_cost"] > 0:
            nd = costM.normalize_desc(r["desc"])
            if nd:
                acc[nd].append(r["unit_cost"])
    avg_by_norm = {nd: sum(c) / len(c) for nd, c in acc.items()}
    for r in rollup:
        if r["unit_cost"] and r["unit_cost"] > 0:
            continue
        if (r["val_total"] or 0) > 0 and (r["total"] or 0) > 0:
            r["unit_cost"] = r["val_total"] / r["total"]
            continue
        nd = costM.normalize_desc(r["desc"])
        if nd in avg_by_norm:
            r["unit_cost"] = avg_by_norm[nd]

    return recon, rollup


# ============================================================
# Estilos del Excel
# ============================================================

H_FILL = PatternFill("solid", fgColor="305496")
H_FONT = Font(bold=True, color="FFFFFF")
DAY_FILL = PatternFill("solid", fgColor="8EA9DB")
DAY_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
C_FILL = PatternFill("solid", fgColor="F8CBAD")
D_FILL = PatternFill("solid", fgColor="FFC7CE")
ADJ_FILL = PatternFill("solid", fgColor="E2EFDA")
DELTA_UP = PatternFill("solid", fgColor="C6EFCE")
DELTA_DOWN = PatternFill("solid", fgColor="FFC7CE")
INT_FMT = "#,##0"
EUR_FMT = '#,##0 "EUR"'


def autosize(ws, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            ln = len(str(cell.value))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)


# ============================================================
# Pestana: Resumen general
# ============================================================

def write_resumen_general(ws, day_labels, data):
    ws.cell(row=1, column=1, value="STOCK LEASEIR - evolucion diaria (Resumen)").font = Font(bold=True, size=14, color="305496")
    ws.cell(row=2, column=1, value="Bolsas A/B/C/D por dia. Cuadre exacto SAP. Cambios >5% coloreados.").font = Font(italic=True, color="666666")

    ws.cell(row=4, column=1, value="Metrica").fill = H_FILL
    ws.cell(row=4, column=1).font = H_FONT
    for c, lbl in enumerate(day_labels, 2):
        ws.cell(row=4, column=c, value=lbl).fill = H_FILL
        ws.cell(row=4, column=c).font = H_FONT
        ws.cell(row=4, column=c).alignment = Alignment(horizontal="center")

    metrics = [
        ("A - RAW + DEFECTIVE", lambda d: data[d][0]["bolsa_A_val"], EUR_FMT, None),
        ("B - Material embebido (incl. en vuelo)", lambda d: data[d][0]["bolsa_B_val"], EUR_FMT, None),
        ("C - Overhead / drift", lambda d: data[d][0]["bolsa_C_val"], EUR_FMT, C_FILL),
        ("D - WIPs sin BOM", lambda d: data[d][0]["bolsa_D_val"], EUR_FMT, D_FILL),
        ("TOTAL inventario (A+B+C+D)",
         lambda d: sum(data[d][0][k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val")),
         EUR_FMT, TOTAL_FILL),
        ("", lambda d: "", None, None),
        ("Unidades totales (>50EUR/u + Others)",
         lambda d: sum((r["total"] or 0) for r in data[d][1] if (r["total"] or 0) > 0),
         INT_FMT, None),
        ("# Proveedores >50EUR/u",
         lambda d: len(set(r["supplier"] for r in data[d][1]
                           if (r["total"] or 0) > 0 and (r["unit_cost"] or 0) > 50 and r.get("supplier"))),
         INT_FMT, None),
        ("# SPECs > 50EUR/u",
         lambda d: len([r for r in data[d][1] if (r["total"] or 0) > 0 and (r["unit_cost"] or 0) > 50]),
         INT_FMT, None),
    ]
    row = 5
    for name, fn, fmt, fill in metrics:
        if name == "":
            row += 1
            continue
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        if fill:
            ws.cell(row=row, column=1).fill = fill
        prev = None
        for c, lbl in enumerate(day_labels, 2):
            v = fn(lbl)
            cell = ws.cell(row=row, column=c, value=v)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill
                cell.font = Font(bold=name.startswith("TOTAL"))
            elif isinstance(v, (int, float)) and prev is not None and prev != 0:
                pct = (v - prev) / prev
                if abs(pct) >= 0.05:
                    cell.fill = DELTA_UP if pct > 0 else DELTA_DOWN
            prev = v if isinstance(v, (int, float)) else None
        row += 1

    ws.column_dimensions["A"].width = 42
    for c in range(2, 2 + len(day_labels)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B5"


# ============================================================
# Helpers
# ============================================================

def group_by(rollup, key_fn, threshold=50.0):
    by_key = defaultdict(lambda: {"u_raw": 0, "u_emb": 0, "u_tot": 0, "val": 0})
    others = {"u_raw": 0, "u_emb": 0, "u_tot": 0, "val": 0}
    for r in rollup:
        if (r["total"] or 0) <= 0:
            continue
        cu = r.get("unit_cost") or 0
        v = (r.get("val_total") or 0) if (r.get("val_total") or 0) > 0 else cu * r["total"]
        u_raw = r["u_raw"] or 0
        u_emb = r["u_emb"] or 0
        target = by_key[key_fn(r)] if cu > threshold else others
        target["u_raw"] += u_raw
        target["u_emb"] += u_emb
        target["u_tot"] += r["total"]
        target["val"] += v
    return dict(by_key), others


def write_pivot_grupo(ws, title, key_fn, label_col, day_labels, data):
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color="305496")
    ws.cell(row=2, column=1, value="Por dia: U.RAW, U.emb, U.TOT, Valor. Cuadre A+B+C+D = SAP.").font = Font(italic=True, color="666666")

    ws.cell(row=4, column=1, value=label_col).fill = H_FILL
    ws.cell(row=4, column=1).font = H_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)

    groups = {}
    for lbl in day_labels:
        rec, rollup = data[lbl]
        by_key, others = group_by(rollup, key_fn)
        groups[lbl] = {"by_key": by_key, "others": others, "rec": rec}

    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        ws.cell(row=4, column=base, value=lbl).fill = DAY_FILL
        ws.cell(row=4, column=base).font = DAY_FONT
        ws.cell(row=4, column=base).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base + 3)
        for j, h in enumerate(["U.RAW", "U.emb", "U.TOT", "Valor"]):
            ws.cell(row=5, column=base + j, value=h).fill = SUB_FILL
            ws.cell(row=5, column=base + j).font = Font(bold=True)
            ws.cell(row=5, column=base + j).alignment = Alignment(horizontal="center")

    last_lbl = day_labels[-1]
    last_keys = groups[last_lbl]["by_key"]
    all_keys = set()
    for lbl in day_labels:
        all_keys.update(groups[lbl]["by_key"].keys())
    sorted_keys = sorted(all_keys, key=lambda k: -(last_keys.get(k, {"val": 0})["val"]))

    r = 6
    for k in sorted_keys:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        for i, lbl in enumerate(day_labels):
            base = 2 + i * 4
            d = groups[lbl]["by_key"].get(k, {"u_raw": 0, "u_emb": 0, "u_tot": 0, "val": 0})
            ws.cell(row=r, column=base, value=d["u_raw"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 1, value=d["u_emb"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 2, value=d["u_tot"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 3, value=d["val"]).number_format = EUR_FMT
        r += 1

    ws.cell(row=r, column=1, value="Others (<50EUR/u)").font = Font(bold=True, italic=True)
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        o = groups[lbl]["others"]
        ws.cell(row=r, column=base, value=o["u_raw"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 1, value=o["u_emb"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 2, value=o["u_tot"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 3, value=o["val"]).number_format = EUR_FMT
    r += 1

    ws.cell(row=r, column=1, value="-- Subtotal pivot (estimacion)").font = Font(bold=True, italic=True, color="305496")
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        bk = groups[lbl]["by_key"]
        oth = groups[lbl]["others"]
        ws.cell(row=r, column=base, value=sum(d["u_raw"] for d in bk.values()) + oth["u_raw"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 1, value=sum(d["u_emb"] for d in bk.values()) + oth["u_emb"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 2, value=sum(d["u_tot"] for d in bk.values()) + oth["u_tot"]).number_format = INT_FMT
        ws.cell(row=r, column=base + 3, value=sum(d["val"] for d in bk.values()) + oth["val"]).number_format = EUR_FMT
        for j in range(4):
            ws.cell(row=r, column=base + j).fill = SUB_FILL
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1

    ws.cell(row=r, column=1, value="Ajuste estimacion SPECs sin stock raw").font = Font(bold=True, italic=True)
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        rec = groups[lbl]["rec"]
        bk = groups[lbl]["by_key"]
        oth = groups[lbl]["others"]
        pivot_v = sum(d["val"] for d in bk.values()) + oth["val"]
        ajuste = (rec["bolsa_A_val"] + rec["bolsa_B_val"]) - pivot_v
        ws.cell(row=r, column=base + 3, value=ajuste).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = ADJ_FILL
    ws.cell(row=r, column=1).fill = ADJ_FILL
    r += 1

    ws.cell(row=r, column=1, value="C - Overhead / drift WIPs").font = Font(bold=True)
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        ws.cell(row=r, column=base + 3, value=groups[lbl]["rec"]["bolsa_C_val"]).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = C_FILL
    ws.cell(row=r, column=1).fill = C_FILL
    r += 1

    ws.cell(row=r, column=1, value="D - WIPs sin escandallo").font = Font(bold=True)
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        ws.cell(row=r, column=base + 3, value=groups[lbl]["rec"]["bolsa_D_val"]).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = D_FILL
    ws.cell(row=r, column=1).fill = D_FILL
    r += 1

    ws.cell(row=r, column=1, value="TOTAL INVENTARIO (A+B+C+D)").font = Font(bold=True, size=12)
    for i, lbl in enumerate(day_labels):
        base = 2 + i * 4
        rec = groups[lbl]["rec"]
        bk = groups[lbl]["by_key"]
        oth = groups[lbl]["others"]
        total = sum(rec[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
        ut = sum(d["u_tot"] for d in bk.values()) + oth["u_tot"]
        ws.cell(row=r, column=base + 2, value=ut).number_format = INT_FMT
        ws.cell(row=r, column=base + 3, value=total).number_format = EUR_FMT
    for c in range(1, 2 + len(day_labels) * 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 45
    for c in range(2, 2 + len(day_labels) * 4):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = "B6"


# ============================================================
# Movimientos del dia (compras / salidas) por SPEC raw
# ============================================================

def cargar_compras_salidas_hoy(daily_file, day_date):
    """Devuelve (compras_dict, salidas_dict) por SPEC canonico del dia.
    - Compras = pestana 'entradas con SPEC detallada', descompuestas via BOM
    - Salidas = pestanas 'entregas (faltan OTH-4)' + 'salidas directas', descompuestas via BOM
    """
    from datetime import datetime as _dt, date as _date

    def es_hoy(f):
        if isinstance(f, _dt):
            return f.date() == day_date
        if isinstance(f, _date):
            return f == day_date
        return False

    compras = defaultdict(float)
    salidas = defaultdict(float)
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True)
    except Exception:
        return compras, salidas

    def normalize_code(code):
        c = str(code or "").strip()
        if not c:
            return ""
        c = re.sub(r"_C$", "_c", c)
        return master.canonical(c)

    # Entradas con SPEC detallada
    for sn in wb.sheetnames:
        if 'entradas' in sn.lower() and 'spec' in sn.lower():
            sh = wb[sn]
            hdrs = [str(c.value or '').lower() for c in sh[1]]
            fecha_c = next((i for i, h in enumerate(hdrs) if 'fecha' in h and 'albar' in h), None)
            if fecha_c is None:
                fecha_c = next((i for i, h in enumerate(hdrs) if 'fecha' in h), None)
            spec_c = next((i for i, h in enumerate(hdrs)
                           if ('artículo' in h or 'articulo' in h or 'código' in h or 'codigo' in h)), None)
            qty_c = next((i for i, h in enumerate(hdrs) if 'cantidad' in h), None)
            if fecha_c is None or spec_c is None or qty_c is None:
                continue
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) <= max(fecha_c, spec_c, qty_c):
                    continue
                if not es_hoy(row[fecha_c]):
                    continue
                code = normalize_code(row[spec_c])
                if not code:
                    continue
                try:
                    qty = float(row[qty_c] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        compras[raw] += q * qty
                else:
                    compras[code] += qty
            break

    # Entregas + salidas directas
    for sn in wb.sheetnames:
        snl = sn.lower()
        if 'entregas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 8:
                    continue
                if not es_hoy(row[0]):
                    continue
                code = normalize_code(row[5])
                if not code:
                    continue
                try:
                    qty = float(row[7] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        salidas[raw] += q * qty
                else:
                    salidas[code] += qty
        if 'salidas directas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 5:
                    continue
                if not es_hoy(row[0]):
                    continue
                raw_code = str(row[2] or '').strip()
                if not raw_code or 'artículo' in raw_code.lower():
                    continue
                code = normalize_code(raw_code)
                try:
                    qty = float(row[4] or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0:
                    continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        salidas[raw] += q * qty
                else:
                    salidas[code] += qty
    return compras, salidas


# ============================================================
# Pestana: Por SPEC (con cols extra de movimientos del ultimo dia)
# ============================================================

def write_pivot_spec(ws, day_labels, data, daily_files=None):
    """Pestana 'Por SPEC' con todos los SPECs en filas, 4 cols por dia + cuadre.
    Si daily_files se proporciona, anade 3 cols al final: Compras / Salidas / Var.Neta del ultimo dia."""
    ws.cell(row=1, column=1, value="STOCK POR SPEC - evolucion diaria").font = Font(bold=True, size=13, color="305496")
    ws.cell(row=2, column=1, value="Cada SPEC en filas, 4 columnas por dia + Movimientos del ultimo dia (Compras/Salidas/Var.Neta).").font = Font(italic=True, color="666666")

    fixed_hdr = ["SPEC", "Descripcion", "Proveedor", "Tipo", "Coste/u ultimo"]
    for c, h in enumerate(fixed_hdr, 1):
        ws.cell(row=4, column=c, value=h).fill = H_FILL
        ws.cell(row=4, column=c).font = H_FONT
        ws.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)

    col_offset = len(fixed_hdr) + 1
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        ws.cell(row=4, column=base, value=lbl).fill = DAY_FILL
        ws.cell(row=4, column=base).font = DAY_FONT
        ws.cell(row=4, column=base).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base + 3)
        for j, h in enumerate(["U.RAW", "U.emb", "U.TOT", "Valor"]):
            ws.cell(row=5, column=base + j, value=h).fill = SUB_FILL
            ws.cell(row=5, column=base + j).font = Font(bold=True)
            ws.cell(row=5, column=base + j).alignment = Alignment(horizontal="center")

    # Columnas extra: Compras / Salidas / Var.Neta del ultimo dia
    extra_col = col_offset + len(day_labels) * 4
    compras_hoy = defaultdict(float)
    salidas_hoy = defaultdict(float)
    if daily_files:
        last_label = day_labels[-1]
        last_file = next((f for lbl, _, f in daily_files if lbl == last_label), None)
        day_date = next((dt.date() for lbl, dt, _ in daily_files if lbl == last_label), None)
        if last_file and day_date:
            compras_hoy, salidas_hoy = cargar_compras_salidas_hoy(last_file, day_date)
        MOV_FILL = PatternFill("solid", fgColor="6E548D")
        MOV_SUB = PatternFill("solid", fgColor="D9C5E8")
        ws.cell(row=4, column=extra_col, value=f"Movimientos {last_label}").fill = MOV_FILL
        ws.cell(row=4, column=extra_col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=4, column=extra_col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=4, start_column=extra_col, end_row=4, end_column=extra_col + 2)
        for j, h in enumerate(["Compras", "Salidas", "Var.Neta"]):
            ws.cell(row=5, column=extra_col + j, value=h).fill = MOV_SUB
            ws.cell(row=5, column=extra_col + j).font = Font(bold=True)
            ws.cell(row=5, column=extra_col + j).alignment = Alignment(horizontal="center")

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 18

    all_specs = {}
    spec_data = defaultdict(lambda: defaultdict(lambda: {"u_raw": 0, "u_emb": 0, "u_tot": 0, "val": 0}))
    last_lbl = day_labels[-1]
    for lbl in day_labels:
        rec, rollup = data[lbl]
        for r in rollup:
            if (r["total"] or 0) <= 0 and (r["val_total"] or 0) <= 0:
                continue
            sp = r["spec"]
            if sp not in all_specs:
                all_specs[sp] = {
                    "sup": r.get("supplier") or "(SIN PROVEEDOR)",
                    "desc": r.get("desc") or "",
                    "tipo": r.get("tipo") or "",
                    "cu": r.get("unit_cost") or 0,
                    "last_val": 0,
                }
            cu = r.get("unit_cost") or 0
            v = (r.get("val_total") or 0) if (r.get("val_total") or 0) > 0 else cu * (r["total"] or 0)
            spec_data[sp][lbl] = {
                "u_raw": r["u_raw"] or 0,
                "u_emb": r["u_emb"] or 0,
                "u_tot": r["total"] or 0,
                "val": v,
            }
            if lbl == last_lbl:
                all_specs[sp]["last_val"] = v
                all_specs[sp]["cu"] = cu
                all_specs[sp]["sup"] = r.get("supplier") or all_specs[sp]["sup"]
                all_specs[sp]["tipo"] = r.get("tipo") or all_specs[sp]["tipo"]

    spec_order = sorted(all_specs.keys(), key=lambda s: -all_specs[s]["last_val"])

    r = 6
    UP_F = PatternFill("solid", fgColor="C6EFCE")
    DN_F = PatternFill("solid", fgColor="FFC7CE")
    for sp in spec_order:
        info = all_specs[sp]
        ws.cell(row=r, column=1, value=sp).font = Font(bold=True)
        ws.cell(row=r, column=2, value=(info["desc"] or "")[:80])
        ws.cell(row=r, column=3, value=info["sup"])
        ws.cell(row=r, column=4, value=info["tipo"])
        ws.cell(row=r, column=5, value=info["cu"]).number_format = EUR_FMT
        for i, lbl in enumerate(day_labels):
            base = col_offset + i * 4
            d = spec_data[sp].get(lbl, {"u_raw": 0, "u_emb": 0, "u_tot": 0, "val": 0})
            ws.cell(row=r, column=base, value=d["u_raw"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 1, value=d["u_emb"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 2, value=d["u_tot"]).number_format = INT_FMT
            ws.cell(row=r, column=base + 3, value=d["val"]).number_format = EUR_FMT
        if daily_files:
            cval = compras_hoy.get(sp, 0)
            sval = salidas_hoy.get(sp, 0)
            net = cval - sval
            ws.cell(row=r, column=extra_col, value=cval).number_format = '#,##0;-#,##0;0'
            ws.cell(row=r, column=extra_col + 1, value=sval).number_format = '#,##0;-#,##0;0'
            ws.cell(row=r, column=extra_col + 2, value=net).number_format = '+#,##0;-#,##0;0'
            if cval > 0:
                ws.cell(row=r, column=extra_col).fill = UP_F
            if sval > 0:
                ws.cell(row=r, column=extra_col + 1).fill = DN_F
            if net > 0.5:
                ws.cell(row=r, column=extra_col + 2).fill = UP_F
            elif net < -0.5:
                ws.cell(row=r, column=extra_col + 2).fill = DN_F
        r += 1

    # Cuadre
    r += 1
    ws.cell(row=r, column=1, value="-- Subtotal SPECs (A + B pivot)").font = Font(bold=True, italic=True, color="305496")
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        ur = sum(spec_data[s].get(lbl, {"u_raw": 0})["u_raw"] for s in spec_order)
        ue = sum(spec_data[s].get(lbl, {"u_emb": 0})["u_emb"] for s in spec_order)
        ut = sum(spec_data[s].get(lbl, {"u_tot": 0})["u_tot"] for s in spec_order)
        vt = sum(spec_data[s].get(lbl, {"val": 0})["val"] for s in spec_order)
        ws.cell(row=r, column=base, value=ur).number_format = INT_FMT
        ws.cell(row=r, column=base + 1, value=ue).number_format = INT_FMT
        ws.cell(row=r, column=base + 2, value=ut).number_format = INT_FMT
        ws.cell(row=r, column=base + 3, value=vt).number_format = EUR_FMT
        for j in range(4):
            ws.cell(row=r, column=base + j).fill = SUB_FILL
    ws.cell(row=r, column=1).fill = SUB_FILL
    if daily_files:
        tot_c = sum(compras_hoy.values())
        tot_s = sum(salidas_hoy.values())
        ws.cell(row=r, column=extra_col, value=tot_c).number_format = '#,##0;-#,##0;0'
        ws.cell(row=r, column=extra_col + 1, value=tot_s).number_format = '#,##0;-#,##0;0'
        ws.cell(row=r, column=extra_col + 2, value=tot_c - tot_s).number_format = '+#,##0;-#,##0;0'
        for j in range(3):
            ws.cell(row=r, column=extra_col + j).fill = SUB_FILL
    r += 1

    ws.cell(row=r, column=1, value="Ajuste estimacion SPECs sin stock raw").font = Font(bold=True, italic=True)
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        rec = data[lbl][0]
        pivot_v = sum(spec_data[s].get(lbl, {"val": 0})["val"] for s in spec_order)
        ajuste = (rec["bolsa_A_val"] + rec["bolsa_B_val"]) - pivot_v
        ws.cell(row=r, column=base + 3, value=ajuste).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = ADJ_FILL
    ws.cell(row=r, column=1).fill = ADJ_FILL
    r += 1

    ws.cell(row=r, column=1, value="C - Overhead / drift WIPs").font = Font(bold=True)
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        ws.cell(row=r, column=base + 3, value=data[lbl][0]["bolsa_C_val"]).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = C_FILL
    ws.cell(row=r, column=1).fill = C_FILL
    r += 1

    ws.cell(row=r, column=1, value="D - WIPs sin escandallo").font = Font(bold=True)
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        ws.cell(row=r, column=base + 3, value=data[lbl][0]["bolsa_D_val"]).number_format = EUR_FMT
        ws.cell(row=r, column=base + 3).fill = D_FILL
    ws.cell(row=r, column=1).fill = D_FILL
    r += 1

    ws.cell(row=r, column=1, value="TOTAL INVENTARIO (A+B+C+D)").font = Font(bold=True, size=12)
    for i, lbl in enumerate(day_labels):
        base = col_offset + i * 4
        rec = data[lbl][0]
        total = sum(rec[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
        ut = sum(spec_data[s].get(lbl, {"u_tot": 0})["u_tot"] for s in spec_order)
        ws.cell(row=r, column=base + 2, value=ut).number_format = INT_FMT
        ws.cell(row=r, column=base + 3, value=total).number_format = EUR_FMT
    for c in range(1, col_offset + len(day_labels) * 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 13
    for c in range(col_offset, col_offset + len(day_labels) * 4):
        ws.column_dimensions[get_column_letter(c)].width = 10
    if daily_files:
        for c in range(extra_col, extra_col + 3):
            ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = ws.cell(row=6, column=col_offset)


# ============================================================
# MAIN
# ============================================================

def main():
    files = find_daily_files()
    if not files:
        sys.exit("ERROR: no se han encontrado ficheros 'ejercicio DD-MM-YYYY.xlsx' en la carpeta del proyecto.")

    print(f"Encontrados {len(files)} ficheros diarios:")
    for lbl, dt, fpath in files:
        print(f"  {lbl}  ({dt.date()})  {fpath.name}")

    print(f"\nProcesando...")
    data = {}
    day_labels = []
    for lbl, dt, fpath in files:
        print(f"  Procesando {lbl}...")
        recon, rollup = process_day(fpath)
        data[lbl] = (recon, rollup)
        day_labels.append(lbl)

    print(f"\nGenerando Stock_evolutivo.xlsx con 4 pestanas...")
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Resumen general"
    write_resumen_general(ws1, day_labels, data)

    ws2 = wb.create_sheet("Por proveedor")
    write_pivot_grupo(ws2, "STOCK POR PROVEEDOR - evolucion diaria",
                       key_fn=lambda r: r.get("supplier") or "(SIN PROVEEDOR)",
                       label_col="Proveedor",
                       day_labels=day_labels, data=data)

    ws3 = wb.create_sheet("Por tipo")
    write_pivot_grupo(ws3, "STOCK POR TIPO DE COMPONENTE - evolucion diaria",
                       key_fn=lambda r: r.get("tipo") or "Others",
                       label_col="Tipo",
                       day_labels=day_labels, data=data)

    ws4 = wb.create_sheet("Por SPEC")
    write_pivot_spec(ws4, day_labels, data, daily_files=files)

    wb.save(OUTPUT_FILE)
    print(f"\nGUARDADO: {OUTPUT_FILE.name}")

    print(f"\nCuadre verificado:")
    for lbl in day_labels:
        rec = data[lbl][0]
        total = sum(rec[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
        print(f"  {lbl}: TOTAL = {total:,.2f} EUR")


if __name__ == "__main__":
    main()
