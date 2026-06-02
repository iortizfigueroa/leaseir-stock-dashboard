#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stock_por_proveedor.py
======================

Genera Stock_por_proveedor.xlsx con DOS pestanas:
  1. Por proveedor (coste): outline > SPECs ordenados por coste/u descendente
  2. Por tipo de componente: outline > SPECs ordenados por valor

Aplica reclasificacion fina para evitar mezclar (ej. pedal guard ya no es
water pump, capillaries chrome-coated van a Mechanized en vez de Optics).
"""

from __future__ import annotations
import sys
import re
import importlib.util
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parent if "__file__" in dir() else Path(".")
PROJECT_DIR = Path(r"/sessions/stoic-gallant-albattani/mnt/Donet (inventario)")

DAILY_FILE = PROJECT_DIR / "ejercicio 28-05-2026.xlsx"
OUTPUT_FILE = PROJECT_DIR / "Stock_por_proveedor.xlsx"

# Importar inventario_master reutilizando si ya esta cargado por el que nos llama.
# Esto evita el bug de que cada modulo tenga su propia instancia con ESCANDALLOS={}
sys.path.insert(0, str(PROJECT_DIR))
if "master" in sys.modules:
    master = sys.modules["master"]
else:
    spec_m = importlib.util.spec_from_file_location("master", PROJECT_DIR / "inventario_master.py")
    master = importlib.util.module_from_spec(spec_m)
    spec_m.loader.exec_module(master)
    sys.modules["master"] = master
# Si los escandallos no estan cargados, cargarlos
if not master.ESCANDALLOS:
    master.load_escandallos()


def refine_type(tipo, supplier, desc, unit_cost):
    """Reclasifica items mal etiquetados por keyword matching."""
    sup = (supplier or "").upper()
    d = (desc or "")

    # === OVERRIDES PRIORITARIOS (cualquier tipo inicial) ===
    # Hailea HC500A son chillers reales -> Refrigerator
    if re.search(r"hailea|HC500A|chiller cart|\bnevera\b", d, re.IGNORECASE):
        return "Refrigerator"
    # Iwaki Water Pump (descripcion "Water pump" tal cual)
    if re.search(r"^water pump$|\biwaki\b.*pump|water pump\s*-", d, re.IGNORECASE):
        return "Water pump"

    if tipo == "Water pump":
        # Descartar por descripcion ANTES de aprobar por supplier
        # (IWAKI tiene mas SPECs aparte de bombas)
        if re.search(r"\bpedal\b|\bguard\b", d, re.IGNORECASE):
            return "Others"
        if re.search(r"\bvalve\b|\brelay\b|\bsolenoid\b", d, re.IGNORECASE):
            return "Electronic components"
        if re.search(r"\bvacuum\b|\bhydraulic\b", d, re.IGNORECASE):
            return "Others"
        # Ahora aprobar por IWAKI si pasa las negativas
        if "IWAKI" in sup:
            return "Water pump"
        # Y si no, exigir keyword explicito en descripcion
        if not re.search(r"\bwater pump\b|\bmicro pump\b|\bgear pump\b|\bpiston pump\b|\bbomba\b", d, re.IGNORECASE):
            return "Others"

    if tipo == "Power Supply":
        if re.search(r"^WIR-|\bWIR-\d|\bharness\b|\bwire\b|\bcable\b", d, re.IGNORECASE):
            return "Wires"

    if tipo == "Computer":
        if re.search(r"\bsupport\b|\bholder\b|\bcarcasa\b", d, re.IGNORECASE) and not re.search(r"touchscreen|computer", d, re.IGNORECASE):
            return "Mechanized components"

    if tipo == "Refrigerator":
        # PRIMERO: descartar por keywords negativas (mismo patron que water pump)
        if re.search(r"transistor|capacitor|relay|switch|\bconnector\b|enable|cooling unit", d, re.IGNORECASE):
            return "Electronic components"
        if re.search(r"\bCA AC\b|\bCA CHILLER\b", d, re.IGNORECASE):
            return "Electronic components"
        if re.search(r"\bfilter\b", d, re.IGNORECASE):
            return "Mechanized components"
        if re.search(r"\bcover\b|\bfitting\b|\bmount\b|\bspacer\b|\bdisc\b|\bbase\b|\btray\b", d, re.IGNORECASE):
            return "Mechanized components"
        # DESPUES: aprobar refrigerator real
        if re.search(r"hailea|HC500A|chiller cart|nevera", d, re.IGNORECASE):
            return "Refrigerator"
        if "UNIVATEQ" in sup and re.search(r"chiller|cooling|fan|refriger", d, re.IGNORECASE):
            return "Refrigerator"
        if re.search(r"\bfan\b", d, re.IGNORECASE):
            return "Refrigerator"
        return "Others"

    if tipo == "Optics/sapphires/prisms":
        if re.search(r"\bcapillary\b|\bchannel", d, re.IGNORECASE):
            return "Mechanized components"
        if re.search(r"\badhesive\b|norland", d, re.IGNORECASE):
            return "Others"

    if tipo == "Diode":
        if "MONOCROM" not in sup and not re.search(r"\bENSPSTM\b|ensptm|\bLBS-\b|\blaser diode\b", d, re.IGNORECASE):
            return "Electronic components"

    return tipo


def load_inventory_from_xlsx(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sh = wb["inventario"]
    rows = []
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or not r[0]:
            continue
        code = str(r[0]).strip()
        desc = str(r[1] or "").strip()
        try:
            qty = float(r[2]) if r[2] is not None else 0.0
        except (TypeError, ValueError):
            qty = 0.0
        try:
            val = float(r[3]) if r[3] is not None else 0.0
        except (TypeError, ValueError):
            val = 0.0
        if qty <= 0 and val <= 0:
            continue
        rows.append({
            "code": code, "canon": master.canonical(code), "base": master.base_spec(code),
            "desc": desc, "qty": qty, "val_inv": val,
            "unit_cost_inv": (val / qty) if qty > 0 else 0.0, "source": "inv",
        })
    return rows


def load_ofs_from_xlsx(xlsx_path: Path, inventory):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sh = wb["en vuelo"]
    cost_by_canon = {}
    for row in inventory:
        if row["unit_cost_inv"] > 0:
            cost_by_canon[row["canon"]] = row["unit_cost_inv"]
            cost_by_canon.setdefault(row["base"], row["unit_cost_inv"])

    def cost_material_teorico(canon_spec):
        if canon_spec not in master.ESCANDALLOS: return 0.0
        exp = master.expand_bom(canon_spec)
        total = 0.0
        for raw_canon, q in exp.items():
            uc = cost_by_canon.get(raw_canon) or cost_by_canon.get(master.base_spec(raw_canon)) or 0.0
            total += q * uc
        return total

    rows = []
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or r[0] is None: continue
        of_num = r[0]
        spec_code = (str(r[2]) if r[2] is not None else "").strip()
        desc = (str(r[3]) if r[3] is not None else "").strip()
        emitido = float(r[5]) if r[5] is not None else 0.0
        recibido = float(r[6]) if r[6] is not None else 0.0
        pendiente = float(r[7]) if r[7] is not None else 0.0
        if not spec_code or pendiente <= 0: continue
        spec_clean = re.sub(r"_C$", "_c", spec_code.strip())
        canon = master.canonical(spec_clean)
        coste_u = cost_by_canon.get(canon) or cost_by_canon.get(master.base_spec(canon)) or 0.0
        if coste_u == 0.0:
            coste_u = cost_material_teorico(canon)
        if coste_u > 0:
            qty_total = round(emitido / coste_u) if emitido > 0 else 1
            qty_recibida = round(recibido / coste_u) if recibido > 0 else 0
            qty_pendiente = max(0, qty_total - qty_recibida) or 1
        else:
            qty_pendiente = 1
        rows.append({
            "code": f"OF-{of_num}/{spec_clean}", "canon": canon, "base": master.base_spec(canon),
            "desc": f"[EN VUELO OF {of_num}] {desc}", "qty": float(qty_pendiente), "val_inv": pendiente,
            "unit_cost_inv": pendiente / qty_pendiente if qty_pendiente > 0 else 0.0, "source": "of",
        })
    return rows


def load_suppliers_from_xlsx(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sh = wb["proveedores-SPEC"]
    sup_set = defaultdict(set)
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or not r[0]: continue
        code = str(r[0]).strip()
        sup_name = str(r[3] or "").strip()
        if sup_name:
            sup_set[master.canonical(code)].add(sup_name)
    suppliers = {}
    for canon, names in sup_set.items():
        suppliers[canon] = next(iter(names)) if len(names) == 1 else " / ".join(sorted(names))
    return suppliers


def main():
    print("[1/6] Cargando escandallos...")
    master.load_escandallos()
    print(f"      {len(master.ESCANDALLOS)} BOMs")

    print("[2/6] Cargando proveedores y datos del dia...")
    suppliers = load_suppliers_from_xlsx(DAILY_FILE)
    inventory = load_inventory_from_xlsx(DAILY_FILE)
    of_rows = load_ofs_from_xlsx(DAILY_FILE, inventory)
    inventory.extend(of_rows)
    total_inv = sum(r["val_inv"] for r in inventory)
    print(f"      {len(inventory)} items, total = {total_inv:,.2f} EUR ({len(of_rows)} OFs en vuelo)")

    print("[3/6] Clasificando...")
    for row in inventory:
        row["cat"] = master.classify(row, suppliers)

    print("[4/6] Reconciliando bolsas...")
    unit_cost = master.build_unit_costs(inventory)
    recon = master.reconcile(inventory, unit_cost)
    print(f"      A={recon['bolsa_A_val']:,.2f}  B={recon['bolsa_B_val']:,.2f}  C={recon['bolsa_C_val']:,.2f}  D={recon['bolsa_D_val']:,.2f}")

    print("[5/6] Rollup + reclasificacion fina...")
    rollup_rows = master.build_rollup(inventory, recon, unit_cost, suppliers, {})

    # APLICAR refine_type
    n_changed = 0
    for r in rollup_rows:
        old = r["tipo"]
        r["tipo"] = refine_type(old, r.get("supplier", ""), r.get("desc", ""), r.get("unit_cost", 0))
        if r["tipo"] != old:
            n_changed += 1
    print(f"      {n_changed} SPECs reclasificados")

    rows = [r for r in rollup_rows if (r["total"] or 0) > 0 or (r["val_total"] or 0) > 0]

    print("[6/6] Escribiendo Excel...")
    write_excel(rows)
    print(f"GUARDADO: {OUTPUT_FILE}")


def write_excel(rows):
    wb = openpyxl.Workbook()

    H_FILL = PatternFill("solid", fgColor="305496")
    H_FONT = Font(bold=True, color="FFFFFF")
    SUP_FILL = PatternFill("solid", fgColor="8EA9DB")
    SUP_FONT = Font(bold=True, size=12)
    TIPO_FILL = PatternFill("solid", fgColor="FCE4D6")
    TIPO_FONT = Font(bold=True, size=11)
    TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
    EUR = '#,##0.00 "EUR"'
    INT = '#,##0'
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_hdr(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = H_FILL; cell.font = H_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # =========== PESTANA 1: Por proveedor (ordenado por coste/u DESC) ===========
    ws = wb.active
    ws.title = "Por proveedor (coste)"
    write_hdr(ws, 1, ["Proveedor / SPEC", "Descripcion", "Tipo", "Coste/u EUR",
                       "U. RAW", "U. emb", "U. TOTAL", "Valor TOTAL"])
    ws.row_dimensions[1].height = 30

    by_sup = defaultdict(list)
    for r in rows:
        by_sup[r["supplier"] or "(SIN PROVEEDOR)"].append(r)
    sup_order = sorted(by_sup.keys(), key=lambda s: -sum(r["val_total"] for r in by_sup[s]))

    row = 2
    for sup in sup_order:
        items = by_sup[sup]
        n = len(items); ur = sum(r["u_raw"] for r in items); ue = sum(r["u_emb"] for r in items)
        v = sum(r["val_total"] for r in items)
        ws.cell(row=row, column=1, value=f"{sup}  ({n} SPECs)")
        ws.cell(row=row, column=5, value=ur).number_format = INT
        ws.cell(row=row, column=6, value=ue).number_format = INT
        ws.cell(row=row, column=7, value=ur+ue).number_format = INT
        ws.cell(row=row, column=8, value=v).number_format = EUR
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = SUP_FILL
            ws.cell(row=row, column=c).font = SUP_FONT
            ws.cell(row=row, column=c).border = border
        row += 1
        for it in sorted(items, key=lambda x: -(x["unit_cost"] or 0)):
            ws.cell(row=row, column=1, value=f'   {it["spec"]}')
            ws.cell(row=row, column=2, value=(it["desc"] or "")[:90])
            ws.cell(row=row, column=3, value=it["tipo"])
            ws.cell(row=row, column=4, value=it["unit_cost"]).number_format = EUR
            ws.cell(row=row, column=5, value=it["u_raw"]).number_format = INT
            ws.cell(row=row, column=6, value=it["u_emb"]).number_format = INT
            ws.cell(row=row, column=7, value=it["total"]).number_format = INT
            ws.cell(row=row, column=8, value=it["val_total"]).number_format = EUR
            for c in range(1, 9): ws.cell(row=row, column=c).border = border
            ws.row_dimensions[row].outline_level = 1
            row += 1
    for i, w in enumerate([50, 65, 22, 13, 12, 12, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = False

    # =========== PESTANA 2: Por tipo de componente ===========
    ws2 = wb.create_sheet("Por tipo de componente")
    write_hdr(ws2, 1, ["Tipo / SPEC", "Descripcion", "Proveedor", "Coste/u EUR",
                       "U. RAW", "U. emb", "U. TOTAL", "Valor TOTAL"])
    ws2.row_dimensions[1].height = 30

    by_tipo = defaultdict(list)
    for r in rows:
        by_tipo[r["tipo"] or "Others"].append(r)
    tipo_order = sorted(by_tipo.keys(), key=lambda t: -sum(r["val_total"] for r in by_tipo[t]))

    row = 2
    for tipo in tipo_order:
        items = by_tipo[tipo]
        n = len(items); ur = sum(r["u_raw"] for r in items); ue = sum(r["u_emb"] for r in items)
        v = sum(r["val_total"] for r in items)
        ws2.cell(row=row, column=1, value=f"{tipo}  ({n} SPECs)")
        ws2.cell(row=row, column=5, value=ur).number_format = INT
        ws2.cell(row=row, column=6, value=ue).number_format = INT
        ws2.cell(row=row, column=7, value=ur+ue).number_format = INT
        ws2.cell(row=row, column=8, value=v).number_format = EUR
        for c in range(1, 9):
            ws2.cell(row=row, column=c).fill = TIPO_FILL
            ws2.cell(row=row, column=c).font = TIPO_FONT
            ws2.cell(row=row, column=c).border = border
        row += 1
        for it in sorted(items, key=lambda x: -(x["unit_cost"] or 0)):
            ws2.cell(row=row, column=1, value=f'   {it["spec"]}')
            ws2.cell(row=row, column=2, value=(it["desc"] or "")[:90])
            ws2.cell(row=row, column=3, value=it["supplier"] or "(SIN PROVEEDOR)")
            ws2.cell(row=row, column=4, value=it["unit_cost"]).number_format = EUR
            ws2.cell(row=row, column=5, value=it["u_raw"]).number_format = INT
            ws2.cell(row=row, column=6, value=it["u_emb"]).number_format = INT
            ws2.cell(row=row, column=7, value=it["total"]).number_format = INT
            ws2.cell(row=row, column=8, value=it["val_total"]).number_format = EUR
            for c in range(1, 9): ws2.cell(row=row, column=c).border = border
            ws2.row_dimensions[row].outline_level = 1
            row += 1
    for i, w in enumerate([45, 65, 40, 13, 12, 12, 12, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.sheet_properties.outlinePr.summaryBelow = False

    # =========== PESTANA 0: Resumen tipos ===========
    ws3 = wb.create_sheet("Resumen tipos", 0)
    ws3["A1"] = "STOCK POR TIPO DE COMPONENTE - al 27-05-2026"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = "Tipos refinados: pump solo bombas reales, refrigerator solo Hailea/Univateq, etc."
    ws3["A3"].font = Font(italic=True, color="666666")
    write_hdr(ws3, 5, ["Tipo", "# SPECs", "U. RAW", "U. emb", "U. TOTAL", "Coste/u medio", "Valor TOTAL"])
    row = 6
    sum_total = 0
    for tipo in tipo_order:
        items = by_tipo[tipo]
        n = len(items); ur = sum(r["u_raw"] for r in items); ue = sum(r["u_emb"] for r in items)
        t = ur + ue; v = sum(r["val_total"] for r in items)
        avg = v / t if t > 0 else 0
        sum_total += v
        ws3.cell(row=row, column=1, value=tipo).font = Font(bold=True)
        ws3.cell(row=row, column=2, value=n).number_format = INT
        ws3.cell(row=row, column=3, value=ur).number_format = INT
        ws3.cell(row=row, column=4, value=ue).number_format = INT
        ws3.cell(row=row, column=5, value=t).number_format = INT
        ws3.cell(row=row, column=6, value=avg).number_format = EUR
        ws3.cell(row=row, column=7, value=v).number_format = EUR
        for c in range(1, 8): ws3.cell(row=row, column=c).border = border
        row += 1
    ws3.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=12)
    ws3.cell(row=row, column=7, value=sum_total).number_format = EUR
    for c in range(1, 8):
        ws3.cell(row=row, column=c).fill = TOTAL_FILL
        ws3.cell(row=row, column=c).font = Font(bold=True, size=12)
        ws3.cell(row=row, column=c).border = border
    for i, w in enumerate([35, 10, 14, 14, 14, 16, 18], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)

if __name__ == "__main__":
    main()
