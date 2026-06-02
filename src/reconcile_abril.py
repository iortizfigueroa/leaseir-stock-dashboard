#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
reconcile_abril.py
==================

Partiendo del inventario del 29-05-2026 y aplicando hacia atras todas las
entradas y salidas del mes (entre 4-mayo y 29-mayo), reconstruye el
inventario del 30-04-2026 esperado y lo compara con el real.

Logica al nivel de TOTAL (raw + embebido):
  total_30abr[spec] = total_29may[spec] - sum(entradas) + sum(salidas)

donde:
  - sum(entradas) = compras del mes a nivel raw (descompuestas via BOM)
  - sum(salidas) = entregas WIPs (descompuestas a raws) + salidas directas

Salidas como SUMARLAS porque vamos al reves en el tiempo.
"""

from __future__ import annotations
import sys
import re
import importlib.util
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parent
ABRIL_FILE = PROJECT_DIR / "Inventario 30 abril.xlsx"
MAYO_FILE = PROJECT_DIR / "ejercicio 29-05-2026.xlsx"
OUTPUT_FILE = PROJECT_DIR / "Reconcile_abril_via_mayo.xlsx"

sys.path.insert(0, str(PROJECT_DIR))

spec_m = importlib.util.spec_from_file_location("evol", PROJECT_DIR / "actualizar_evolutivo.py")
evol = importlib.util.module_from_spec(spec_m)
spec_m.loader.exec_module(evol)
master = evol.master
prov = evol.prov


def cargar_movimientos_periodo(daily_file):
    """Igual que cargar_compras_salidas_hoy pero sin filtrar por fecha (todo el rango)."""
    compras = defaultdict(float)
    salidas = defaultdict(float)
    try:
        wb = openpyxl.load_workbook(daily_file, data_only=True)
    except Exception:
        return compras, salidas

    def norm(c):
        c = str(c or "").strip()
        if not c: return ""
        c = re.sub(r"_C$", "_c", c)
        return master.canonical(c)

    for sn in wb.sheetnames:
        if 'entradas' in sn.lower() and 'spec' in sn.lower():
            sh = wb[sn]
            hdrs = [str(c.value or '').lower() for c in sh[1]]
            spec_c = next((i for i, h in enumerate(hdrs)
                           if ('artículo' in h or 'articulo' in h or 'código' in h or 'codigo' in h)), None)
            qty_c = next((i for i, h in enumerate(hdrs) if 'cantidad' in h), None)
            if spec_c is None or qty_c is None:
                continue
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) <= max(spec_c, qty_c):
                    continue
                code = norm(row[spec_c])
                if not code: continue
                try: qty = float(row[qty_c] or 0)
                except: continue
                if qty <= 0: continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        compras[raw] += q * qty
                else:
                    compras[code] += qty
            break

    for sn in wb.sheetnames:
        snl = sn.lower()
        if 'entregas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 8 or not row[5]: continue
                code = norm(row[5])
                if not code: continue
                try: qty = float(row[7] or 0)
                except: continue
                if qty <= 0: continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        salidas[raw] += q * qty
                else:
                    salidas[code] += qty
        if 'salidas directas' in snl:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True, min_row=2):
                if not row or len(row) < 5: continue
                raw_code = str(row[2] or '').strip()
                if not raw_code or 'artículo' in raw_code.lower(): continue
                code = norm(raw_code)
                try: qty = float(row[4] or 0)
                except: continue
                if qty <= 0: continue
                if code in master.ESCANDALLOS:
                    for raw, q in master.expand_bom(code).items():
                        salidas[raw] += q * qty
                else:
                    salidas[code] += qty
    return compras, salidas


def build_totales_por_spec(daily_file):
    """Devuelve {spec: {u_tot, u_raw, u_emb, val, unit_cost, desc, supplier, tipo}}."""
    recon, rollup, inv, uc = evol.process_day_full(daily_file) if hasattr(evol, "process_day_full") else _fallback(daily_file)
    # Inferir Monocrom + refine_type
    for r in rollup:
        if not r.get("supplier"):
            d = r.get("desc", "") or ""
            if re.search(r"laser diode|\bENSPSTM\b|ensptm|\bLBS-|\bmonocrom\b", d, re.IGNORECASE):
                r["supplier"] = "MONOCROM S.L"
        r["tipo"] = prov.refine_type(r["tipo"], r.get("supplier", ""), r.get("desc", ""), r.get("unit_cost", 0))
    info = {}
    for r in rollup:
        sp = r["spec"]
        info[sp] = {
            "u_tot": r["total"] or 0,
            "u_raw": r["u_raw"] or 0,
            "u_emb": r["u_emb"] or 0,
            "val": r["val_total"] or 0,
            "unit_cost": r["unit_cost"] or 0,
            "desc": r["desc"] or "",
            "supplier": r["supplier"] or "(SIN PROVEEDOR)",
            "tipo": r["tipo"] or "",
        }
    return info, recon


def _load_inventory_robust(xlsx_path):
    """Carga inventario detectando layout (mayuscula minuscula, columnas extra)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Encontrar sheet por nombre case-insensitive
    inv_sheet = None
    for sn in wb.sheetnames:
        if sn.lower().strip() == "inventario":
            inv_sheet = sn
            break
    if inv_sheet is None:
        raise KeyError("No hay sheet 'inventario' en " + str(xlsx_path))
    sh = wb[inv_sheet]
    # Detectar columnas via headers
    hdrs = [str(c.value or '').lower().strip() for c in sh[1]]
    code_c = next((i for i, h in enumerate(hdrs) if 'artículo' in h or 'articulo' in h or 'número' in h or 'numero' in h or 'code' in h or h == 'spec' or 'código' in h or 'codigo' in h), None)
    desc_c = next((i for i, h in enumerate(hdrs) if 'descripción' in h or 'descripcion' in h or 'description' in h), None)
    qty_c = next((i for i, h in enumerate(hdrs) if 'cantidad' in h or 'qty' in h), None)
    val_c = next((i for i, h in enumerate(hdrs) if 'valor acum' in h or 'valor' in h and 'busc' not in h), None)
    if code_c is None or qty_c is None:
        raise ValueError(f"Cols no detectadas en {inv_sheet}: {hdrs}")
    rows = []
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or len(r) <= max(code_c, qty_c) or not r[code_c]:
            continue
        code = str(r[code_c]).strip()
        if not code or code.lower() == "número de artículo":
            continue
        desc = str(r[desc_c] or "").strip() if desc_c is not None else ""
        try: qty = float(r[qty_c] or 0)
        except: qty = 0.0
        try: val = float(r[val_c] or 0) if val_c is not None else 0.0
        except: val = 0.0
        if qty <= 0 and val <= 0:
            continue
        rows.append({
            "code": code, "canon": master.canonical(code), "base": master.base_spec(code),
            "desc": desc, "qty": qty, "val_inv": val,
            "unit_cost_inv": (val / qty) if qty > 0 else 0.0, "source": "inv",
        })
    return rows


def _load_ofs_robust(daily_file, inv):
    """Carga OFs en vuelo con nombre de sheet case-insensitive."""
    import openpyxl as _xl
    wb = _xl.load_workbook(daily_file, data_only=True)
    of_sheet = None
    for sn in wb.sheetnames:
        if 'vuelo' in sn.lower():
            of_sheet = sn
            break
    if of_sheet is None:
        return []
    # Monkey-patch temporal: copiar sheet a nombre 'en vuelo' y delegar
    # Mas robusto: replicar la logica aqui usando el sheet correcto
    sh = wb[of_sheet]
    cost_by_canon = {}
    for row in inv:
        if row.get("unit_cost_inv", 0) > 0:
            cost_by_canon[row["canon"]] = row["unit_cost_inv"]
            cost_by_canon.setdefault(row["base"], row["unit_cost_inv"])

    def cost_material_teorico(canon_spec):
        if canon_spec not in master.ESCANDALLOS: return 0.0
        exp = master.expand_bom(canon_spec)
        total = 0.0
        for raw_canon, q in exp.items():
            uc = cost_by_canon.get(raw_canon) or cost_by_canon.get(master.base_spec(raw_canon)) or 0.0
            total += q * uc
        return total

    rows = []
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or r[0] is None: continue
        of_num = r[0]
        spec_code = (str(r[2]) if r[2] is not None else "").strip()
        desc = (str(r[3]) if r[3] is not None else "").strip()
        try: emitido = float(r[5]) if r[5] is not None else 0.0
        except: emitido = 0.0
        try: recibido = float(r[6]) if r[6] is not None else 0.0
        except: recibido = 0.0
        try: pendiente = float(r[7]) if r[7] is not None else 0.0
        except: pendiente = 0.0
        if not spec_code or pendiente <= 0: continue
        spec_clean = re.sub(r"_C$", "_c", spec_code.strip())
        canon = master.canonical(spec_clean)
        coste_u = cost_by_canon.get(canon) or cost_by_canon.get(master.base_spec(canon)) or 0.0
        if coste_u == 0.0:
            coste_u = cost_material_teorico(canon)
        if coste_u > 0:
            qty_total = round(emitido / coste_u) if emitido > 0 else 1
            qty_recibida = round(recibido / coste_u) if recibido > 0 else 0
            qty_pendiente = max(0, qty_total - qty_recibida) or 1
        else:
            qty_pendiente = 1
        rows.append({
            "code": f"OF-{of_num}/{spec_clean}", "canon": canon, "base": master.base_spec(canon),
            "desc": f"[EN VUELO OF {of_num}] {desc}", "qty": float(qty_pendiente), "val_inv": pendiente,
            "unit_cost_inv": pendiente / qty_pendiente if qty_pendiente > 0 else 0.0, "source": "of",
        })
    return rows


def _fallback(daily_file):
    try:
        suppliers = prov.load_suppliers_from_xlsx(daily_file)
    except Exception:
        suppliers = {}
    try:
        inv = prov.load_inventory_from_xlsx(daily_file)
    except KeyError:
        inv = _load_inventory_robust(daily_file)
    try:
        of_rows = prov.load_ofs_from_xlsx(daily_file, inv)
    except Exception:
        of_rows = _load_ofs_robust(daily_file, inv)
    inv.extend(of_rows)
    for row in inv:
        row["cat"] = master.classify(row, suppliers)
    unit_cost = master.build_unit_costs(inv)
    recon = master.reconcile(inv, unit_cost)
    rollup = master.build_rollup(inv, recon, unit_cost, suppliers, {})
    return recon, rollup, inv, unit_cost


def main():
    print("[1/4] Cargando 29-mayo (origen)...")
    info_29, recon_29 = build_totales_por_spec(MAYO_FILE)
    print(f"  {len(info_29)} SPECs en 29-mayo")

    print("[2/4] Cargando 30-abril (real)...")
    info_30, recon_30 = build_totales_por_spec(ABRIL_FILE)
    print(f"  {len(info_30)} SPECs en 30-abril")

    print("[3/4] Cargando movimientos del periodo (4-mayo / 29-mayo)...")
    compras_total, salidas_total = cargar_movimientos_periodo(MAYO_FILE)
    print(f"  Compras: {sum(compras_total.values()):,.0f} uds en {len(compras_total)} SPECs")
    print(f"  Salidas: {sum(salidas_total.values()):,.0f} uds en {len(salidas_total)} SPECs")

    print("[4/4] Calculando 30-abril inferido = 29-mayo - compras + salidas...")
    all_specs = set(info_29.keys()) | set(info_30.keys()) | set(compras_total.keys()) | set(salidas_total.keys())

    rows = []
    for sp in all_specs:
        i29 = info_29.get(sp, {"u_tot": 0, "unit_cost": 0, "desc": "", "supplier": "(SIN PROVEEDOR)", "tipo": ""})
        i30 = info_30.get(sp, {"u_tot": 0, "unit_cost": 0, "desc": "", "supplier": "(SIN PROVEEDOR)", "tipo": ""})
        c = compras_total.get(sp, 0)
        s = salidas_total.get(sp, 0)
        inferred = i29["u_tot"] - c + s
        real = i30["u_tot"]
        diff = inferred - real
        cu = i29["unit_cost"] or i30["unit_cost"] or 0
        rows.append({
            "spec": sp,
            "desc": (i29["desc"] or i30["desc"])[:80],
            "supplier": i29["supplier"] or i30["supplier"],
            "tipo": i29["tipo"] or i30["tipo"],
            "cu": cu,
            "u_29may": i29["u_tot"],
            "compras": c,
            "salidas": s,
            "inferred_30abr": inferred,
            "real_30abr": real,
            "diff_uds": diff,
            "diff_eur": diff * cu,
            "abs_diff_eur": abs(diff * cu),
        })

    rows.sort(key=lambda r: -r["abs_diff_eur"])

    total_inferred_eur = sum(r["inferred_30abr"] * r["cu"] for r in rows)
    total_real_eur = sum(r["real_30abr"] * r["cu"] for r in rows)
    total_diff_uds = sum(r["diff_uds"] for r in rows)
    total_diff_eur = sum(r["diff_eur"] for r in rows)
    total_abs_diff_eur = sum(r["abs_diff_eur"] for r in rows)

    real_total_30abr = sum(recon_30[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
    real_total_29may = sum(recon_29[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))

    print()
    print("=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"Inventario REAL 30-abril:   {real_total_30abr:>15,.2f} EUR  (oficial via 4-bolsas)")
    print(f"Inventario REAL 29-mayo:    {real_total_29may:>15,.2f} EUR  (oficial via 4-bolsas)")
    print(f"Inferido total 30-abril:    {total_inferred_eur:>15,.2f} EUR  (sumando rollup x coste)")
    print(f"Real    total 30-abril:     {total_real_eur:>15,.2f} EUR  (sumando rollup x coste)")
    print(f"Diff neto (signo importa):  {total_diff_eur:>+15,.2f} EUR  ({total_diff_uds:+,.0f} uds)")
    print(f"Diff absoluto (suma |x|):   {total_abs_diff_eur:>15,.2f} EUR")
    print()
    print("TOP 20 SPECs con mayor discrepancia:")
    print(f"{'SPEC':<14}{'desc':<40}{'29may':>7}{'-c':>7}{'+s':>7}{'inf':>7}{'real':>7}{'diff':>7}{'EUR':>13}")
    for r in rows[:20]:
        print(f"{r['spec']:<14}{r['desc'][:38]:<40}"
              f"{r['u_29may']:>7.0f}{-r['compras']:>+7.0f}{r['salidas']:>+7.0f}"
              f"{r['inferred_30abr']:>7.0f}{r['real_30abr']:>7.0f}"
              f"{r['diff_uds']:>+7.0f}{r['diff_eur']:>+12,.0f}")

    print()
    print(f"Generando Excel: {OUTPUT_FILE.name}...")
    write_excel(rows, recon_29, recon_30, compras_total, salidas_total)
    print(f"GUARDADO: {OUTPUT_FILE.name}")


def write_excel(rows, recon_29, recon_30, compras_total, salidas_total):
    wb = openpyxl.Workbook()

    H_FILL = PatternFill("solid", fgColor="305496")
    H_FONT = Font(bold=True, color="FFFFFF")
    SUB_FILL = PatternFill("solid", fgColor="BDD7EE")
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")
    BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
    NEUT_FILL = PatternFill("solid", fgColor="FFEB9C")
    TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
    INT = "#,##0"
    INT_SIGNED = "+#,##0;-#,##0;0"
    EUR = '#,##0.00 "EUR"'
    EUR_SIGNED = '+#,##0.00 "EUR";-#,##0.00 "EUR";0'

    # === Pestania 1: Resumen ===
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value="RECONCILIACION 30-abril desde 29-mayo via movimientos del mes").font = Font(bold=True, size=14, color="305496")
    ws.cell(row=2, column=1, value="Aplica entradas (-) y salidas (+) del periodo 4/5-29/5 sobre el inventario del 29-may para inferir el 30-abr").font = Font(italic=True, color="666666")

    real_30 = sum(recon_30[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
    real_29 = sum(recon_29[k] for k in ("bolsa_A_val", "bolsa_B_val", "bolsa_C_val", "bolsa_D_val"))
    inferred_total_eur = sum(r["inferred_30abr"] * r["cu"] for r in rows)
    real_total_eur = sum(r["real_30abr"] * r["cu"] for r in rows)
    diff_uds = sum(r["diff_uds"] for r in rows)
    diff_eur = sum(r["diff_eur"] for r in rows)
    abs_diff_eur = sum(r["abs_diff_eur"] for r in rows)

    r = 4
    pairs = [
        ("Inventario REAL 30-abril (SAP)", real_30, EUR, OK_FILL),
        ("Inventario REAL 29-mayo (SAP)", real_29, EUR, OK_FILL),
        ("Movimiento real abril -> mayo", real_29 - real_30, EUR_SIGNED, NEUT_FILL),
        ("", "", None, None),
        ("Total inferido 30-abril (sum spec x cu)", inferred_total_eur, EUR, None),
        ("Total real 30-abril (sum spec x cu)", real_total_eur, EUR, None),
        ("Diff neta (inferido - real)", diff_eur, EUR_SIGNED, BAD_FILL if abs(diff_eur) > 1000 else OK_FILL),
        ("Diff absoluta (suma |x|)", abs_diff_eur, EUR, None),
        ("", "", None, None),
        ("Suma compras periodo", sum(compras_total.values()), INT, None),
        ("Suma salidas periodo", sum(salidas_total.values()), INT, None),
        ("Diff uds netas (inferido - real)", diff_uds, INT_SIGNED, None),
        ("# SPECs analizados", len(rows), INT, None),
        ("# SPECs con discrepancia >0.5 uds", sum(1 for x in rows if abs(x["diff_uds"]) > 0.5), INT, None),
    ]
    for name, val, fmt, fill in pairs:
        if name == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=val)
        if fmt: c.number_format = fmt
        if fill:
            ws.cell(row=r, column=1).fill = fill
            c.fill = fill
        r += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22

    # === Pestania 2: Detalle por SPEC ===
    ws2 = wb.create_sheet("Detalle por SPEC")
    hdr = ["#", "SPEC", "Descripcion", "Proveedor", "Tipo", "Coste/u",
           "29-may uds", "(-) Compras", "(+) Salidas", "Inferido 30-abr", "Real 30-abr",
           "Diff uds", "Diff EUR"]
    for c_, h in enumerate(hdr, 1):
        ws2.cell(row=1, column=c_, value=h).fill = H_FILL
        ws2.cell(row=1, column=c_).font = H_FONT
        ws2.cell(row=1, column=c_).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(rows, start=2):
        diff_uds_v = row["diff_uds"]
        diff_eur_v = row["diff_eur"]
        ws2.cell(row=i, column=1, value=i-1).number_format = INT
        ws2.cell(row=i, column=2, value=row["spec"])
        ws2.cell(row=i, column=3, value=row["desc"])
        ws2.cell(row=i, column=4, value=row["supplier"])
        ws2.cell(row=i, column=5, value=row["tipo"])
        ws2.cell(row=i, column=6, value=row["cu"]).number_format = EUR
        ws2.cell(row=i, column=7, value=row["u_29may"]).number_format = INT
        ws2.cell(row=i, column=8, value=-row["compras"]).number_format = INT_SIGNED
        ws2.cell(row=i, column=9, value=row["salidas"]).number_format = INT_SIGNED
        ws2.cell(row=i, column=10, value=row["inferred_30abr"]).number_format = INT
        ws2.cell(row=i, column=11, value=row["real_30abr"]).number_format = INT
        ws2.cell(row=i, column=12, value=diff_uds_v).number_format = INT_SIGNED
        ws2.cell(row=i, column=13, value=diff_eur_v).number_format = EUR_SIGNED
        if abs(diff_eur_v) > 100:
            for c_ in range(1, 14):
                ws2.cell(row=i, column=c_).fill = BAD_FILL
        elif abs(diff_eur_v) > 10:
            for c_ in range(1, 14):
                ws2.cell(row=i, column=c_).fill = NEUT_FILL
        elif abs(diff_uds_v) < 0.5:
            ws2.cell(row=i, column=12).fill = OK_FILL
            ws2.cell(row=i, column=13).fill = OK_FILL

    widths = [5, 14, 50, 30, 25, 12, 12, 12, 12, 14, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "C2"
    ws2.row_dimensions[1].height = 32

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    main()
