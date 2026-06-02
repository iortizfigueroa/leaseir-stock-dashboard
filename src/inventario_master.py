#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inventario_master.py
====================

Procesa el inventario diario de Leaseir (CSV exportado de SAP B1) y produce
un Excel maestro con varias vistas:

  - Resumen         : totales por bolsa (A/B/C/D), categoria, y tipo
  - Inventario      : 1022 items tal cual del CSV (suma = 2,208,692.22 EUR)
  - Rollup_por_RAW  : cada SPEC raw -> unidades fisicas en toda la cadena
                      (en stock raw + embebidas en WIP + embebidas en End Product)
  - Bolsa_A_raw     : RAW + DEFECTIVE en stock
  - Bolsa_BC_WIP    : WIPs/EPs con escandallo, con su materia (B) y drift (C)
  - Bolsa_D_WIP     : WIPs/EPs sin escandallo
  - SPECs_BOM       : todos los escandallos cargados, formato padre + hijos

Las assertions al final verifican que TODO cuadra a +/- TOLERANCE_EUR antes
de guardar. Si no cuadra, lanza error y NO escribe el Excel.

Uso:
    python3 inventario_master.py
    python3 inventario_master.py path/al/inventario.csv path/al/output.xlsx

CONFIG: toca las constantes abajo si:
  - Aparece un SPEC nuevo en BOM y no en inventario  -> mete proxy en PROXY_COSTS
  - Aparece un WIP nuevo sin escandallo formal       -> mete su receta en SYNTHETIC_BOMS
  - Cambia el supplier de Qualio -> tipo de componente  -> SUPPLIER_TO_TYPE
"""

from __future__ import annotations
import csv
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================

# Directorio donde vive este script. Todo se busca relativo a aqui.
PROJECT_DIR = Path(__file__).resolve().parent

DEFAULT_INPUT_CSV = PROJECT_DIR / "Inventario 30 Abril .csv"
DEFAULT_OUTPUT_XLSX = PROJECT_DIR / "Inventario_MASTER.xlsx"
QUALIO_SUPPLIERS_CSV = PROJECT_DIR / "Qualio_RAW_fabricantes.csv"

# Fichero opcional con OFs en vuelo (Ordenes de Fabricacion en curso).
# Pestania 2 = "OF's en vuelo". Columnas: OF | Status | SPEC | Desc | Fecha | Emitido | Recibido | Pendiente
# Cada OF se anade al inventario como pseudo-item WIP (qty inferida).
OFS_EN_VUELO_XLSX = PROJECT_DIR / "Stock 30 abril.xlsx"
OFS_EN_VUELO_SHEET = "OF´s en vuelo"  # OF[apostrofe ascii alto]s en vuelo

# Carpeta con los escandallos (una subcarpeta por familia con 1+ xlsx).
# El nombre de la hoja Excel ES el SPEC padre (ej. "SPEC-1392").
# Recorre recursivamente todos los xlsx.
SPECS_FOLDER = PROJECT_DIR / "SPECS"

# Fallback: si llegara a faltar la carpeta SPECS, usar estos xlsx en plano
LEGACY_ESCANDALLO_FILES = [
    PROJECT_DIR / "escandallos MHR.xlsx",
    PROJECT_DIR / "escandallos cosmetic.xlsx",
    PROJECT_DIR / "escandallos XCELL y MDD.xlsx",
    PROJECT_DIR / "escandallos AHR, DS y SRF.xlsx",
]

# Tolerancia para las assertions
TOLERANCE_EUR = 1.00

# BOMs sinteticos para WIP que no tienen escandallo formal.
# Formato: SPEC_padre_canonico -> [(SPEC_hijo, cantidad), ...]
# El SPEC_padre_canonico es el codigo SIN barras (canonical()).
SYNTHETIC_BOMS: dict[str, list[tuple[str, float]]] = {
    # Casos legacy que ya cubrian los escandallos antiguos pero pueden faltar
    # en el nuevo set. Si se duplican, el escandallo formal gana.
}

# Coste proxy para SPECs que estan en BOM pero NO en inventario (phantoms).
# Solo cubrimos los Monocrom diodes phantoms (alto valor). El resto valoran 0.
PROXY_COSTS: dict[str, float] = {
    "SPEC-944": 2500.0,  # Monocrom 10 bars 808nm 170A 2000W
    "SPEC-909": 3500.0,  # Monocrom 20 bars 808-940-1060nm 4000W
    "SPEC-86":  2800.0,  # Monocrom LBS-8010-8x8-2-1S
    "SPEC-87":  3500.0,
    "SPEC-92":  3000.0,
    "SPEC-435": 3000.0,
    "SPEC-733": 2200.0,
}

# Mapping proveedor -> tipo de componente (19 categorias)
SUPPLIER_TO_TYPE: dict[str, str] = {
    "MONOCROM SL":              "Diode",
    "DONGGUAN":                 "Mechanized components",
    "RS AMIDATA SAU":           "Electronic components",
    "FARNELL":                  "Electronic components",
    "ARROW IBERIA ELECTRONICA SLU": "Electronic components",
    "DIGI-KEY":                 "Electronic components",
    "RUTRONIK":                 "Electronic components",
    "KOLBI ELECTRONICA SA":     "Power Supply",
    "TDK":                      "Power Supply",
    "MEAN WELL":                "Power Supply",
    "IWAKI EUROPE GMBH":        "Water pump",
    "UNIVATEQ":                 "Refrigerator",
    "TENTE RUEDAS SA":          "Wheels",
    "LASERMET LTD":             "Glasses",
    "HARTING":                  "Harting connectors",
    "COMERCIAL BECANI SL":      "Mechanized components",
    "ROALDO TECNICAS DE FIJACION SL": "Mechanized components",
}

# Fallback por keyword en descripcion si el proveedor no esta mapeado.
DESC_KEYWORDS_TO_TYPE = [
    (r"\bdiode\b|laser diode|monocrom|enspstm|ensptm", "Diode"),
    (r"motherboard|pcba|pcb|main board",               "Motherboard and wires"),
    (r"power supply|fuente|psu|alimentaci",            "Power Supply"),
    (r"computer|pc industrial|industrial pc|touchscreen|advantech", "Computer"),
    (r"pump|bomba",                                    "Water pump"),
    (r"refriger|chiller|nevera|cooling|radiator|fan", "Refrigerator"),
    (r"frontal|cover|casing|cover|handpiece casing",   "Frontal and Handpiece casing"),
    (r"paint|pintura",                                 "Painting"),
    (r"wire|cable|cabledado|harness|umbilical",        "Wires"),
    (r"capacitor|condensad",                           "Condensator"),
    (r"box|caja(?! de )|enclosure",                    "Box"),
    (r"briefcase|maleta",                              "Briefcase"),
    (r"wheel|rueda|caster",                            "Wheels"),
    (r"gafa|glasses|safety glasses",                   "Glasses"),
    (r"harting",                                       "Harting connectors"),
    (r"prism|optic|sapphire|lens|ventana|window|lente", "Optics/sapphires/prisms"),
    (r"screw|tornillo|bolt|nut|washer|arandela|chassis|frame|bracket|mechanized|mecanizad", "Mechanized components"),
]


# ============================================================
# UTILIDADES
# ============================================================

def parse_eur(s) -> float:
    """Convierte '1.234,56' o '1234,56' o '1234.56' a float."""
    if s is None:
        return 0.0
    s = str(s).strip()
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def canonical(code: str) -> str:
    """Quita las barras intermedias del codigo.

    SPEC-752/230/D_c -> SPEC-752230D_c
    SPEC-479/110     -> SPEC-479110
    SPEC-226/Pack    -> SPEC-226Pack
    """
    return (code or "").strip().replace("/", "")


def base_spec(code: str) -> str:
    """Quita sufijos / variantes / _c para obtener un 'codigo de familia base'.

    SPEC-685/DF -> SPEC-685
    SPEC-685_c  -> SPEC-685
    SPEC-752/230/D_c -> SPEC-752
    """
    c = (code or "").strip()
    c = re.sub(r"/.*$", "", c)
    c = re.sub(r"_c$", "", c, flags=re.IGNORECASE)
    return c


# ============================================================
# CARGA DE ESCANDALLOS
# ============================================================

# SPEC_canonico_padre -> [(SPEC_canonico_hijo, qty), ...]
ESCANDALLOS: dict[str, list[tuple[str, float]]] = {}
# SPEC -> descripcion (la primera no vacia que veamos)
SPEC_DESC_FROM_BOM: dict[str, str] = {}


def _parse_workbook(path: Path) -> None:
    """Lee un xlsx y registra todos los escandallos. Cada hoja con nombre
    'SPEC-XXX' es un padre cuyo escandallo es la lista de filas."""
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        sh_clean = sheet_name.strip()
        if not sh_clean.startswith("SPEC-"):
            continue
        parent_canon = canonical(sh_clean)
        sh = wb[sheet_name]
        items: list[tuple[str, float]] = []
        for row in sh.iter_rows(values_only=True, min_row=2):
            if not row or len(row) < 2:
                continue
            cell1 = row[1]
            if cell1 is None:
                continue
            child = str(cell1).strip()
            if not child.startswith("SPEC-"):
                continue
            desc = str(row[2]).strip() if len(row) >= 3 and row[2] else ""
            qty_cell = row[3] if len(row) >= 4 else None
            try:
                qty = float(qty_cell) if qty_cell is not None else 1.0
            except (TypeError, ValueError):
                qty = 1.0
            child_canon = canonical(child)
            items.append((child_canon, qty))
            if desc and child_canon not in SPEC_DESC_FROM_BOM:
                SPEC_DESC_FROM_BOM[child_canon] = desc
        if items:
            ESCANDALLOS[parent_canon] = items


def load_escandallos():
    """Carga todos los escandallos desde SPECS_FOLDER recursivamente.
    Fallback: LEGACY_ESCANDALLO_FILES si SPECS_FOLDER no existe."""
    if SPECS_FOLDER.exists() and SPECS_FOLDER.is_dir():
        for f in sorted(SPECS_FOLDER.rglob("*.xlsx")):
            try:
                _parse_workbook(f)
            except Exception as e:
                print(f"[WARN] Error leyendo {f.name}: {e}", file=sys.stderr)
    else:
        print(f"[WARN] No existe carpeta SPECS, usando legacy files", file=sys.stderr)
        for f in LEGACY_ESCANDALLO_FILES:
            if f.exists():
                try:
                    _parse_workbook(f)
                except Exception as e:
                    print(f"[WARN] Error leyendo {f.name}: {e}", file=sys.stderr)

    # Sinteticos al final (sobreescriben si chocan)
    for parent, recipe in SYNTHETIC_BOMS.items():
        ESCANDALLOS[canonical(parent)] = [(canonical(c), float(q)) for c, q in recipe]


def expand_bom(spec: str, visited: set[str] | None = None) -> dict[str, float]:
    """Expande recursivamente un SPEC a un dict {raw_canonico: qty_por_unidad}.
    Si el spec no tiene escandallo, se considera raw terminal y devuelve {canonical(spec): 1}.
    """
    if visited is None:
        visited = set()
    c = canonical(spec)
    if c in visited:
        return {c: 1.0}
    visited.add(c)
    recipe = ESCANDALLOS.get(c)
    if not recipe:
        visited.discard(c)
        return {c: 1.0}
    acc: dict[str, float] = defaultdict(float)
    for child_canon, qty in recipe:
        sub = expand_bom(child_canon, visited)
        for r, q in sub.items():
            acc[r] += q * qty
    visited.discard(c)
    return dict(acc)


# ============================================================
# CARGA DEL INVENTARIO
# ============================================================

def load_inventory(csv_path: Path) -> list[dict]:
    """Lee el CSV de SAP B1 (Latin-1, ; separado). Devuelve lista de dicts.

    Cada dict tiene:
      code (original con barras)
      canon (sin barras)
      base (canonical sin sufijos)
      desc, qty, val_inv, unit_cost_inv
    """
    rows = []
    with open(csv_path, encoding="latin-1") as f:
        reader = csv.reader(f, delimiter=";")
        next(reader)
        for r in reader:
            if len(r) < 4:
                continue
            code = (r[0] or "").strip()
            desc = (r[1] or "").strip()
            qty = parse_eur(r[2])
            val = parse_eur(r[3])
            if not code:
                continue
            if qty <= 0 and val <= 0:
                continue
            rows.append({
                "code": code,
                "canon": canonical(code),
                "base": base_spec(code),
                "desc": desc,
                "qty": qty,
                "val_inv": val,
                "unit_cost_inv": (val / qty) if qty > 0 else 0.0,
                "source": "inv",  # marca origen: "inv" = inventario CSV, "of" = OF en vuelo
            })
    return rows


def load_ofs_en_vuelo(xlsx_path: Path, inventory: list[dict]) -> list[dict]:
    """Lee la pestania de OFs en vuelo del Excel y devuelve pseudo-items para
    anadir al inventario.

    Para cada OF:
      - Si el SPEC esta en el inventario actual, usamos su coste_unitario_inv
        para inferir qty de la OF: qty_total = round(emitido/coste_u),
        qty_recibida = round(recibido/coste_u), qty_pendiente = qty_total - qty_recibida.
      - Si no esta en inventario, asumimos qty_pendiente = 1.
      - Valor del pseudo-item = "Coste Emitido y no Recibido" (columna pendiente).

    El pseudo-item tiene code = "OF-<num>-<SPEC>" para evitar colision y poder
    listarlo separadamente. Su canon = canonical(SPEC) para que el resto del
    pipeline (clasificacion, BOM expansion, rollup) lo trate como el SPEC normal.
    """
    if not xlsx_path.exists():
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Buscar la pestania por nombre (acepta variantes de apostrofe)
    sheet_name = None
    for sn in wb.sheetnames:
        if "vuelo" in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        return []
    sh = wb[sheet_name]

    # Coste unitario por canon desde el inventario actual
    cost_by_canon: dict[str, float] = {}
    for row in inventory:
        if row["unit_cost_inv"] > 0:
            cost_by_canon[row["canon"]] = row["unit_cost_inv"]
            cost_by_canon.setdefault(row["base"], row["unit_cost_inv"])

    # Helper: coste material teorico de un SPEC con escandallo
    # (expand_bom * coste_unitario de cada raw)
    def cost_material_teorico(canon_spec: str) -> float:
        if canon_spec not in ESCANDALLOS:
            return 0.0
        exp = expand_bom(canon_spec)
        total = 0.0
        for raw_canon, q in exp.items():
            uc = cost_by_canon.get(raw_canon, 0.0)
            if uc == 0.0:
                uc = cost_by_canon.get(base_spec(raw_canon), 0.0)
            total += q * uc
        return total

    rows = []
    for r in sh.iter_rows(values_only=True, min_row=2):
        if not r or r[0] is None:
            continue
        of_num = r[0]
        status = r[1] if len(r) > 1 else ""
        spec_code = (str(r[2]) if len(r) > 2 and r[2] is not None else "").strip()
        desc = (str(r[3]) if len(r) > 3 and r[3] is not None else "").strip()
        emitido = float(r[5]) if len(r) > 5 and r[5] is not None else 0.0
        recibido = float(r[6]) if len(r) > 6 and r[6] is not None else 0.0
        pendiente = float(r[7]) if len(r) > 7 and r[7] is not None else 0.0
        if not spec_code or pendiente <= 0:
            continue

        # SAP a veces escribe "_C" en mayuscula (ej. SPEC-853_C); normalizamos
        spec_clean = re.sub(r"_C$", "_c", spec_code.strip())
        canon = canonical(spec_clean)

        # Inferir cantidad:
        # 1) Si el SPEC tiene coste unitario en el inventario, usarlo
        # 2) Si no, calcular coste material teorico desde el escandallo
        # 3) Fallback: qty = 1
        coste_u = cost_by_canon.get(canon) or cost_by_canon.get(base_spec(canon)) or 0.0
        fuente_coste = "inv" if coste_u > 0 else ""
        if coste_u == 0.0:
            coste_u = cost_material_teorico(canon)
            if coste_u > 0:
                fuente_coste = "bom"

        if coste_u > 0:
            qty_total = round(emitido / coste_u) if emitido > 0 else 1
            qty_recibida = round(recibido / coste_u) if recibido > 0 else 0
            qty_pendiente = max(0, qty_total - qty_recibida)
            if qty_pendiente == 0:
                qty_pendiente = 1
        else:
            qty_pendiente = 1
            fuente_coste = "fallback"

        rows.append({
            "code": f"OF-{of_num}/{spec_clean}",
            "canon": canon,
            "base": base_spec(canon),
            "desc": f"[EN VUELO OF {of_num}] {desc}",
            "qty": float(qty_pendiente),
            "val_inv": pendiente,
            "unit_cost_inv": pendiente / qty_pendiente if qty_pendiente > 0 else 0.0,
            "source": "of",
            "of_num": of_num,
            "of_status": status,
            "of_emitido": emitido,
            "of_recibido": recibido,
        })
    return rows


# ============================================================
# QUALIO SUPPLIERS
# ============================================================

def load_qualio_suppliers() -> tuple[dict[str, str], dict[str, str]]:
    """Devuelve dict {SPEC: supplier} y dict {SPEC: qualio_title}."""
    suppliers: dict[str, str] = {}
    titles: dict[str, str] = {}
    if not QUALIO_SUPPLIERS_CSV.exists():
        return suppliers, titles
    with open(QUALIO_SUPPLIERS_CSV, encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        next(reader)
        for r in reader:
            if len(r) < 4:
                continue
            spec = canonical(r[0])
            title = r[1].strip()
            sup = r[3].strip()
            if not spec:
                continue
            titles[spec] = title
            if sup:
                suppliers[spec] = sup
    return suppliers, titles


def supplier_to_type(supplier: str, desc: str) -> str:
    """Devuelve el tipo de componente. Reglas con prioridad:

    1. Anti-patterns: brochure/manual/label -> Others (incluso si dicen "diode")
    2. Mounts/bases/brackets de cualquier cosa -> Mechanized
    3. Lens/prism/sapphire/window/optic -> Optics (incluso si proveedor = Monocrom)
    4. Diodos electronicos (switching, 1N..., Schottky, zener) -> Electronic components
    5. Diodos laser REALES (Monocrom + "laser diode", ENSPSTM, LBS-) -> Diode
    6. Fallback al mapping de supplier
    7. Fallback al keyword en descripcion
    """
    sup = (supplier or "").upper()
    d = (desc or "")

    # 1. Documentacion / etiquetas
    if re.search(r"\bbrochure\b|\bmanual\b|leaflet|catalog|\blabel\b", d, re.IGNORECASE):
        return "Others"

    # 2. Monocrom prioritario: si supplier es Monocrom Y dice diode/laser/ENSPSTM/LBS-,
    #    es Diode aunque tambien lleve "prism" (caso SPEC-350 diode+Dual Prism).
    if "MONOCROM" in sup and re.search(r"\bdiode\b|laser|\bENSPSTM\b|ensptm|\bLBS-", d, re.IGNORECASE):
        return "Diode"

    # 3. Mounts / bases / brackets aunque digan "diode"
    if re.search(r"\bmount\b|\bbracket\b|\bbase\b|chassis|\bframe\b|\btray\b|\bhanger\b", d, re.IGNORECASE):
        return "Mechanized components"

    # 4. Optics (lens/prism/sapphire/window/capillary/channel)
    if re.search(r"\blens\b|\bprism\b|\bsapphire\b|\boptic|\bwindow\b|\bventana\b|\bcapillary\b|\bchannel", d, re.IGNORECASE):
        return "Optics/sapphires/prisms"

    # 5. Diodos electronicos (switching, 1N..., Schottky, zener) - NO son laser
    if re.search(r"switching diode|standard diode|schottky|\b1N\d+\b|\bDO-?\d|zener|rectifier", d, re.IGNORECASE):
        return "Electronic components"

    # 6. Laser diodes sin proveedor Monocrom (ej. SPEC-91/93 ENSPSTM038/033)
    if re.search(r"\blaser diode\b|\bENSPSTM\b|\bensptm\b|\bLBS-", d, re.IGNORECASE):
        return "Diode"

    # 6. Mapping por supplier
    for key, t in SUPPLIER_TO_TYPE.items():
        if key in sup:
            return t

    # 7. Keywords en descripcion
    for pattern, t in DESC_KEYWORDS_TO_TYPE:
        if re.search(pattern, d, re.IGNORECASE):
            return t
    return "Others"


# ============================================================
# CLASIFICACION
# ============================================================

RX_DF = re.compile(r"/DF$|/DEF$|/OER$", re.IGNORECASE)
RX_C  = re.compile(r"_c$", re.IGNORECASE)

DESC_END_KEYS = re.compile(
    r"\b(MHR|MDD|XCELL|SRF|DS|AHR)[\s-]?(110|220|230)\s*Aesth|"
    r"230-Aesth|110-Aesth",
    re.IGNORECASE
)
DESC_DEF_KEYS = re.compile(r"defectu|defect|/df\b|/oer\b", re.IGNORECASE)

DESC_WIP_STRONG = re.compile(
    r"\bleaseir\b.*aesthetic|"
    r"\bconsola\b|\bconsole\b(?!\s+(modular|back))|"
    r"\bcrate\b|\bin\s+crate\b|"
    r"\blaserhead\b|"
    r"^handpiece\b|\bhandpiece\s+(single|dual|quad)",
    re.IGNORECASE
)

DESC_RAW_STRONG = re.compile(
    r"\bpcba?\b|\bpcb\b|\bboard\b|\bbody\b|\bcover\b|\bfrontal\b|"
    r"\bcoupling\b|\bsocket\b|\bcable\b|\bconnector\b|\bhanger\b|"
    r"\bhardcase\b|\bfoam\b|\bbracket\b|\bchassis\b|\bscrew\b|"
    r"\bcomputer\b|\bhuman\s+machine\s+interface\b|\bhmi\b",
    re.IGNORECASE
)


def classify(row: dict, supplier_map: dict[str, str]) -> str:
    """Devuelve 'RAW', 'WIP', 'END_PRODUCT' o 'DEFECTIVE'."""
    code = row["code"]
    desc = row["desc"] or ""
    canon = row["canon"]

    if RX_DF.search(code) or DESC_DEF_KEYS.search(desc):
        return "DEFECTIVE"

    has_bom = canon in ESCANDALLOS

    if DESC_END_KEYS.search(desc):
        return "END_PRODUCT"

    if has_bom:
        return "WIP"

    if RX_C.search(code):
        return "WIP"

    if DESC_RAW_STRONG.search(desc):
        return "RAW"

    if re.match(r"^(FMAT|FFER|PTS|RAW|MAT)-", code):
        return "RAW"

    base = row["base"]
    if base in supplier_map or canon in supplier_map:
        return "RAW"

    if DESC_WIP_STRONG.search(desc):
        return "WIP"

    return "RAW"


# ============================================================
# COSTES UNITARIOS
# ============================================================

def build_unit_costs(inventory: list[dict]) -> dict[str, float]:
    """Para cada SPEC canonico, calcula coste unitario medio ponderado a partir
    de las filas RAW de ese canonico (excluyendo /DF y /OER)."""
    qty_by_canon: dict[str, float] = defaultdict(float)
    val_by_canon: dict[str, float] = defaultdict(float)
    for row in inventory:
        if row["cat"] != "RAW":
            continue
        if "/" in row["code"]:  # variantes raras no entran en coste base
            continue
        qty_by_canon[row["canon"]] += row["qty"]
        val_by_canon[row["canon"]] += row["val_inv"]

    unit_cost: dict[str, float] = {}
    for c, q in qty_by_canon.items():
        if q > 0:
            unit_cost[c] = val_by_canon[c] / q

    # Tambien por base (sin _c) para fallback
    qty_by_base: dict[str, float] = defaultdict(float)
    val_by_base: dict[str, float] = defaultdict(float)
    for row in inventory:
        if row["cat"] != "RAW":
            continue
        if "/" in row["code"]:
            continue
        b = row["base"]
        qty_by_base[b] += row["qty"]
        val_by_base[b] += row["val_inv"]
    for b, q in qty_by_base.items():
        if q > 0 and b not in unit_cost:
            unit_cost[b] = val_by_base[b] / q

    # Proxies
    for c, cost in PROXY_COSTS.items():
        unit_cost.setdefault(canonical(c), cost)
    return unit_cost


# ============================================================
# RECONCILIACION 4 BOLSAS
# ============================================================

def reconcile(inventory: list[dict], unit_cost: dict[str, float]):
    bolsa_A_val = 0.0
    bolsa_B_val = 0.0
    bolsa_C_val = 0.0
    bolsa_D_val = 0.0

    attrib_A: dict[str, float] = defaultdict(float)
    attrib_B: dict[str, float] = defaultdict(float)
    units_raw: dict[str, float] = defaultdict(float)
    units_emb: dict[str, float] = defaultdict(float)

    wips_b_c: list[dict] = []
    wips_d: list[dict] = []

    for row in inventory:
        cat = row["cat"]
        c = row["canon"]

        if cat in ("RAW", "DEFECTIVE"):
            bolsa_A_val += row["val_inv"]
            attrib_A[c] += row["val_inv"]
            units_raw[c] += row["qty"]
            continue

        # WIP / END_PRODUCT
        if c not in ESCANDALLOS:
            bolsa_D_val += row["val_inv"]
            wips_d.append({
                "code": row["code"], "desc": row["desc"], "cat": cat,
                "qty": row["qty"], "val_inv": row["val_inv"],
            })
            continue

        # Tiene BOM
        bom = expand_bom(c)
        mat_total = 0.0
        per_raw: dict[str, float] = {}
        for raw_canon, qty_per_unit in bom.items():
            uc = unit_cost.get(raw_canon, 0.0)
            if uc == 0.0:
                # Fallback: probar por base sin _c
                bc = base_spec(raw_canon)
                uc = unit_cost.get(bc, 0.0)
            qty_total = qty_per_unit * row["qty"]
            val_contrib = qty_total * uc
            mat_total += val_contrib
            per_raw[raw_canon] = val_contrib
            units_emb[raw_canon] += qty_total

        drift = row["val_inv"] - mat_total
        bolsa_B_val += mat_total
        bolsa_C_val += drift

        for raw_canon, val_contrib in per_raw.items():
            attrib_B[raw_canon] += val_contrib

        wips_b_c.append({
            "code": row["code"], "desc": row["desc"], "cat": cat,
            "qty": row["qty"], "val_inv": row["val_inv"],
            "mat_v": mat_total, "drift": drift,
        })

    return {
        "bolsa_A_val": bolsa_A_val,
        "bolsa_B_val": bolsa_B_val,
        "bolsa_C_val": bolsa_C_val,
        "bolsa_D_val": bolsa_D_val,
        "attrib_A": attrib_A,
        "attrib_B": attrib_B,
        "units_raw": units_raw,
        "units_emb": units_emb,
        "wips_b_c": wips_b_c,
        "wips_d": wips_d,
    }


# ============================================================
# ROLLUP POR RAW
# ============================================================

def build_rollup(inventory, recon, unit_cost, suppliers, titles):
    raw_canons: set[str] = set()

    # 1. Cualquier RAW o DEFECTIVE en inventario
    for row in inventory:
        if row["cat"] in ("RAW", "DEFECTIVE"):
            raw_canons.add(row["canon"])

    # 2. Hojas terminales (sin BOM propio) de cualquier escandallo
    for parent in ESCANDALLOS:
        for raw_canon, _ in [(c, q) for c, q in [(canonical(ch), q) for ch, q in ESCANDALLOS[parent]]]:
            if raw_canon not in ESCANDALLOS:
                raw_canons.add(raw_canon)

    # 3. Tambien recoger hojas de la expansion completa
    for parent in ESCANDALLOS:
        for raw_canon in expand_bom(parent):
            if raw_canon not in ESCANDALLOS:
                raw_canons.add(raw_canon)

    # Descripciones y categoria
    desc_by_canon: dict[str, str] = {}
    cat_by_canon: dict[str, str] = {}
    for row in inventory:
        if row["canon"] not in desc_by_canon or row["cat"] == "RAW":
            desc_by_canon[row["canon"]] = row["desc"]
            cat_by_canon[row["canon"]] = row["cat"]

    rollup_rows = []
    for c in sorted(raw_canons):
        u_raw = recon["units_raw"].get(c, 0.0)
        u_emb = recon["units_emb"].get(c, 0.0)
        total = u_raw + u_emb
        val_A = recon["attrib_A"].get(c, 0.0)
        val_B = recon["attrib_B"].get(c, 0.0)
        uc = unit_cost.get(c, 0.0)
        if uc == 0.0:
            uc = unit_cost.get(base_spec(c), 0.0)
        is_proxy = canonical(c) in {canonical(p) for p in PROXY_COSTS} and (val_A == 0.0)
        desc = desc_by_canon.get(c) or SPEC_DESC_FROM_BOM.get(c) or titles.get(c, "")
        cat = cat_by_canon.get(c, "RAW")
        supplier = suppliers.get(c) or suppliers.get(base_spec(c), "")
        tipo = supplier_to_type(supplier, desc)
        rollup_rows.append({
            "spec": c, "desc": desc, "cat": cat, "tipo": tipo,
            "supplier": supplier,
            "u_raw": u_raw, "u_emb": u_emb, "total": total,
            "unit_cost": uc, "is_proxy": is_proxy,
            "val_A": val_A, "val_B": val_B,
            "val_total": val_A + val_B,
        })
    rollup_rows.sort(key=lambda r: -r["val_total"])
    return rollup_rows


# ============================================================
# EXCEL
# ============================================================

H_FILL = PatternFill("solid", fgColor="305496")
H_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True)
EUR_FMT = '#,##0.00 "EUR"'
INT_FMT = '#,##0'
PCT_FMT = '0.0%'


def _hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_excel(out_path: Path, inventory, recon, rollup_rows, totals):
    wb = openpyxl.Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "INVENTARIO LEASEIR - MAESTRO"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Total: {totals['total_inv']:,.2f} EUR ({totals['n_items']} items) - Fuente: {totals['fecha']}"

    # Bolsas
    ws["A4"] = "RECONCILIACION 4 BOLSAS"
    ws["A4"].font = SUB_FONT
    _hdr(ws, 5, ["Bolsa", "Descripcion", "Valor EUR", "% del total"])
    bolsas = [
        ("A", "RAW + DEFECTIVE en stock", recon["bolsa_A_val"]),
        ("B", "Materiales embebidos en WIP/EP con escandallo", recon["bolsa_B_val"]),
        ("C", "Overhead/drift de precio en WIPs con escandallo", recon["bolsa_C_val"]),
        ("D", "WIP/EP sin escandallo (valor directo)", recon["bolsa_D_val"]),
    ]
    for i, (b, d, v) in enumerate(bolsas, start=6):
        ws.cell(row=i, column=1, value=b).font = SUB_FONT
        ws.cell(row=i, column=2, value=d)
        ws.cell(row=i, column=3, value=v).number_format = EUR_FMT
        ws.cell(row=i, column=4, value=v / totals["total_inv"]).number_format = PCT_FMT
    tot_r = 6 + len(bolsas)
    ws.cell(row=tot_r, column=1, value="TOTAL").font = SUB_FONT
    ws.cell(row=tot_r, column=2, value="Suma A+B+C+D")
    suma = sum(v for _, _, v in bolsas)
    ws.cell(row=tot_r, column=3, value=suma).number_format = EUR_FMT
    ws.cell(row=tot_r, column=3).font = SUB_FONT

    # Categoria
    rs = tot_r + 3
    ws.cell(row=rs, column=1, value="RESUMEN POR CATEGORIA").font = SUB_FONT
    _hdr(ws, rs + 1, ["Categoria", "Items", "Unidades", "Valor EUR", "% valor"])
    cat_summary: dict[str, dict[str, float]] = defaultdict(lambda: {"n": 0, "u": 0.0, "v": 0.0})
    for row in inventory:
        c = cat_summary[row["cat"]]
        c["n"] += 1
        c["u"] += row["qty"]
        c["v"] += row["val_inv"]
    for i, cat in enumerate(["RAW", "WIP", "END_PRODUCT", "DEFECTIVE"], start=rs + 2):
        c = cat_summary.get(cat, {"n": 0, "u": 0.0, "v": 0.0})
        ws.cell(row=i, column=1, value=cat).font = SUB_FONT
        ws.cell(row=i, column=2, value=c["n"]).number_format = INT_FMT
        ws.cell(row=i, column=3, value=c["u"]).number_format = INT_FMT
        ws.cell(row=i, column=4, value=c["v"]).number_format = EUR_FMT
        ws.cell(row=i, column=5, value=c["v"] / totals["total_inv"]).number_format = PCT_FMT

    # Por tipo
    ts = rs + 8
    ws.cell(row=ts, column=1, value="RAW MATERIALS POR TIPO (Bolsa A + B)").font = SUB_FONT
    _hdr(ws, ts + 1, ["Tipo", "Items", "Unidades totales", "Valor A+B EUR", "% A+B"])
    tipo_sum: dict[str, dict[str, float]] = defaultdict(lambda: {"n": 0, "u": 0.0, "v": 0.0})
    for r in rollup_rows:
        t = tipo_sum[r["tipo"]]
        t["n"] += 1
        t["u"] += r["total"]
        t["v"] += r["val_total"]
    AB = recon["bolsa_A_val"] + recon["bolsa_B_val"]
    for i, (tipo, s) in enumerate(sorted(tipo_sum.items(), key=lambda x: -x[1]["v"]), start=ts + 2):
        ws.cell(row=i, column=1, value=tipo)
        ws.cell(row=i, column=2, value=s["n"]).number_format = INT_FMT
        ws.cell(row=i, column=3, value=s["u"]).number_format = INT_FMT
        ws.cell(row=i, column=4, value=s["v"]).number_format = EUR_FMT
        ws.cell(row=i, column=5, value=(s["v"] / AB) if AB else 0).number_format = PCT_FMT
    _autosize(ws)

    # Inventario
    ws = wb.create_sheet("Inventario")
    _hdr(ws, 1, ["Codigo", "Descripcion", "Cantidad", "Valor EUR", "Coste/u EUR", "Categoria", "Canonico", "Tipo", "Proveedor"])
    rollup_by_c = {r["spec"]: r for r in rollup_rows}
    for i, row in enumerate(inventory, start=2):
        rr = rollup_by_c.get(row["canon"])
        supplier = rr["supplier"] if rr else ""
        tipo = rr["tipo"] if rr else ""
        ws.cell(row=i, column=1, value=row["code"])
        ws.cell(row=i, column=2, value=row["desc"])
        ws.cell(row=i, column=3, value=row["qty"]).number_format = '0.00'
        ws.cell(row=i, column=4, value=row["val_inv"]).number_format = EUR_FMT
        ws.cell(row=i, column=5, value=row["unit_cost_inv"]).number_format = EUR_FMT
        ws.cell(row=i, column=6, value=row["cat"])
        ws.cell(row=i, column=7, value=row["canon"])
        ws.cell(row=i, column=8, value=tipo)
        ws.cell(row=i, column=9, value=supplier)
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Rollup
    ws = wb.create_sheet("Rollup_por_RAW")
    _hdr(ws, 1, [
        "SPEC", "Descripcion", "Cat", "Tipo", "Proveedor",
        "U. raw stock", "U. embebidas", "U. TOTALES",
        "Coste/u EUR", "Proxy?",
        "Valor RAW (A)", "Valor embebido (B)", "Valor TOTAL"
    ])
    for i, r in enumerate(rollup_rows, start=2):
        ws.cell(row=i, column=1, value=r["spec"])
        ws.cell(row=i, column=2, value=r["desc"])
        ws.cell(row=i, column=3, value=r["cat"])
        ws.cell(row=i, column=4, value=r["tipo"])
        ws.cell(row=i, column=5, value=r["supplier"])
        ws.cell(row=i, column=6, value=r["u_raw"]).number_format = '0.00'
        ws.cell(row=i, column=7, value=r["u_emb"]).number_format = '0.00'
        ws.cell(row=i, column=8, value=r["total"]).number_format = '0.00'
        ws.cell(row=i, column=9, value=r["unit_cost"]).number_format = EUR_FMT
        ws.cell(row=i, column=10, value="SI" if r["is_proxy"] else "")
        ws.cell(row=i, column=11, value=r["val_A"]).number_format = EUR_FMT
        ws.cell(row=i, column=12, value=r["val_B"]).number_format = EUR_FMT
        ws.cell(row=i, column=13, value=r["val_total"]).number_format = EUR_FMT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Bolsa A
    ws = wb.create_sheet("Bolsa_A_raw_stock")
    _hdr(ws, 1, ["Codigo", "Descripcion", "Cantidad", "Valor EUR", "Categoria", "Tipo", "Proveedor"])
    r_idx = 2
    sA = 0.0
    for row in inventory:
        if row["cat"] in ("RAW", "DEFECTIVE"):
            rr = rollup_by_c.get(row["canon"])
            ws.cell(row=r_idx, column=1, value=row["code"])
            ws.cell(row=r_idx, column=2, value=row["desc"])
            ws.cell(row=r_idx, column=3, value=row["qty"]).number_format = '0.00'
            ws.cell(row=r_idx, column=4, value=row["val_inv"]).number_format = EUR_FMT
            ws.cell(row=r_idx, column=5, value=row["cat"])
            ws.cell(row=r_idx, column=6, value=rr["tipo"] if rr else "")
            ws.cell(row=r_idx, column=7, value=rr["supplier"] if rr else "")
            sA += row["val_inv"]
            r_idx += 1
    ws.cell(row=r_idx + 1, column=3, value="TOTAL A").font = SUB_FONT
    ws.cell(row=r_idx + 1, column=4, value=sA).number_format = EUR_FMT
    ws.cell(row=r_idx + 1, column=4).font = SUB_FONT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Bolsa B+C
    ws = wb.create_sheet("Bolsa_BC_WIP_conBOM")
    _hdr(ws, 1, ["Codigo", "Descripcion", "Categoria", "Stock", "Valor inv", "Materiales B", "Drift/Overhead C"])
    sB = sC = sI = 0.0
    for i, w in enumerate(recon["wips_b_c"], start=2):
        ws.cell(row=i, column=1, value=w["code"])
        ws.cell(row=i, column=2, value=w["desc"])
        ws.cell(row=i, column=3, value=w["cat"])
        ws.cell(row=i, column=4, value=w["qty"]).number_format = '0.00'
        ws.cell(row=i, column=5, value=w["val_inv"]).number_format = EUR_FMT
        ws.cell(row=i, column=6, value=w["mat_v"]).number_format = EUR_FMT
        ws.cell(row=i, column=7, value=w["drift"]).number_format = EUR_FMT
        sI += w["val_inv"]; sB += w["mat_v"]; sC += w["drift"]
    tr = 2 + len(recon["wips_b_c"]) + 1
    ws.cell(row=tr, column=4, value="TOTAL").font = SUB_FONT
    ws.cell(row=tr, column=5, value=sI).number_format = EUR_FMT
    ws.cell(row=tr, column=6, value=sB).number_format = EUR_FMT
    ws.cell(row=tr, column=7, value=sC).number_format = EUR_FMT
    for c in (5, 6, 7):
        ws.cell(row=tr, column=c).font = SUB_FONT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Bolsa D
    ws = wb.create_sheet("Bolsa_D_WIP_sinBOM")
    _hdr(ws, 1, ["Codigo", "Descripcion", "Categoria", "Stock", "Valor EUR"])
    sD = 0.0
    for i, w in enumerate(recon["wips_d"], start=2):
        ws.cell(row=i, column=1, value=w["code"])
        ws.cell(row=i, column=2, value=w["desc"])
        ws.cell(row=i, column=3, value=w["cat"])
        ws.cell(row=i, column=4, value=w["qty"]).number_format = '0.00'
        ws.cell(row=i, column=5, value=w["val_inv"]).number_format = EUR_FMT
        sD += w["val_inv"]
    tr = 2 + len(recon["wips_d"]) + 1
    ws.cell(row=tr, column=4, value="TOTAL").font = SUB_FONT
    ws.cell(row=tr, column=5, value=sD).number_format = EUR_FMT
    ws.cell(row=tr, column=5).font = SUB_FONT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # SPECs BOM
    ws = wb.create_sheet("SPECs_BOM")
    _hdr(ws, 1, ["SPEC padre", "SPEC hijo", "Cantidad", "Desc hijo"])
    i = 2
    for parent in sorted(ESCANDALLOS):
        for child, qty in ESCANDALLOS[parent]:
            ws.cell(row=i, column=1, value=parent)
            ws.cell(row=i, column=2, value=child)
            ws.cell(row=i, column=3, value=qty)
            ws.cell(row=i, column=4, value=SPEC_DESC_FROM_BOM.get(child, ""))
            i += 1
    ws.freeze_panes = "A2"
    _autosize(ws)

    wb.save(out_path)


# ============================================================
# MAIN + ASSERTS
# ============================================================

def main():
    args = sys.argv[1:]
    in_csv = Path(args[0]) if args else DEFAULT_INPUT_CSV
    out_xlsx = Path(args[1]) if len(args) > 1 else DEFAULT_OUTPUT_XLSX

    if not in_csv.exists():
        sys.exit(f"ERROR: no existe el CSV de entrada: {in_csv}")

    print(f"[1/6] Cargando escandallos...")
    load_escandallos()
    print(f"      {len(ESCANDALLOS)} BOMs cargados")

    print(f"[2/6] Cargando Qualio suppliers...")
    suppliers, titles = load_qualio_suppliers()
    print(f"      {len(suppliers)} SPECs con proveedor")

    print(f"[3/6] Cargando inventario: {in_csv.name}")
    inventory = load_inventory(in_csv)
    total_csv = sum(r["val_inv"] for r in inventory)
    print(f"      {len(inventory)} items, total CSV = {total_csv:,.2f} EUR")

    # OFs en vuelo: pseudo-items WIP que se suman al inventario
    if OFS_EN_VUELO_XLSX.exists():
        of_rows = load_ofs_en_vuelo(OFS_EN_VUELO_XLSX, inventory)
        if of_rows:
            val_of = sum(r["val_inv"] for r in of_rows)
            qty_of = sum(r["qty"] for r in of_rows)
            print(f"      + {len(of_rows)} OFs en vuelo: {qty_of:.0f} uds, valor pendiente = {val_of:,.2f} EUR")
            inventory.extend(of_rows)
    total_inv = sum(r["val_inv"] for r in inventory)
    print(f"      TOTAL inventario + en vuelo = {total_inv:,.2f} EUR")

    print(f"[4/6] Clasificando items...")
    for row in inventory:
        row["cat"] = classify(row, suppliers)
    cat_counts = defaultdict(int)
    for row in inventory:
        cat_counts[row["cat"]] += 1
    print(f"      {dict(cat_counts)}")

    print(f"[5/6] Reconciliando 4 bolsas...")
    unit_cost = build_unit_costs(inventory)
    recon = reconcile(inventory, unit_cost)
    print(f"      A = {recon['bolsa_A_val']:>14,.2f}   (RAW + DEFECTIVE)")
    print(f"      B = {recon['bolsa_B_val']:>14,.2f}   (materiales en WIP con BOM)")
    print(f"      C = {recon['bolsa_C_val']:>14,.2f}   (drift/overhead)")
    print(f"      D = {recon['bolsa_D_val']:>14,.2f}   (WIP sin BOM)")
    suma = recon["bolsa_A_val"] + recon["bolsa_B_val"] + recon["bolsa_C_val"] + recon["bolsa_D_val"]
    print(f"      ----------------------------")
    print(f"    SUMA = {suma:>14,.2f}   (inv = {total_inv:,.2f})")

    print(f"[6/6] Construyendo rollup por RAW...")
    rollup_rows = build_rollup(inventory, recon, unit_cost, suppliers, titles)
    print(f"      {len(rollup_rows)} SPECs en rollup")

    # ASSERTS
    print("\n[ASSERTS] Verificando coherencia...")

    diff = abs(suma - total_inv)
    if diff > TOLERANCE_EUR:
        sys.exit(f"ASSERT FAIL: bolsas suman {suma:,.2f} vs inventario {total_inv:,.2f} (diff {diff:,.2f})")
    print(f"      OK Bolsas A+B+C+D = inventario  (diff {diff:.4f} EUR)")

    sum_A = sum(r["val_A"] for r in rollup_rows)
    diff = abs(sum_A - recon["bolsa_A_val"])
    if diff > TOLERANCE_EUR:
        sys.exit(f"ASSERT FAIL: rollup val_A {sum_A:,.2f} vs Bolsa A {recon['bolsa_A_val']:,.2f} (diff {diff:,.2f})")
    print(f"      OK Sum(rollup.val_A) = Bolsa A  (diff {diff:.4f} EUR)")

    sum_B = sum(r["val_B"] for r in rollup_rows)
    diff = abs(sum_B - recon["bolsa_B_val"])
    if diff > TOLERANCE_EUR:
        sys.exit(f"ASSERT FAIL: rollup val_B {sum_B:,.2f} vs Bolsa B {recon['bolsa_B_val']:,.2f} (diff {diff:,.2f})")
    print(f"      OK Sum(rollup.val_B) = Bolsa B  (diff {diff:.4f} EUR)")

    sum_T = sum(r["val_total"] for r in rollup_rows)
    target = recon["bolsa_A_val"] + recon["bolsa_B_val"]
    diff = abs(sum_T - target)
    if diff > TOLERANCE_EUR:
        sys.exit(f"ASSERT FAIL: rollup total {sum_T:,.2f} vs A+B {target:,.2f} (diff {diff:,.2f})")
    print(f"      OK Sum(rollup.val_total) = A+B  (diff {diff:.4f} EUR)")

    print("\n[ASSERTS] Todo OK. Escribiendo Excel...")

    totals = {
        "total_inv": total_inv,
        "n_items": len(inventory),
        "fecha": in_csv.stem,
    }
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_excel(out_xlsx, inventory, recon, rollup_rows, totals)
    print(f"\nGUARDADO: {out_xlsx}")


if __name__ == "__main__":
    main()
